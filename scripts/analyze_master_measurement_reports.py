from __future__ import annotations

import argparse
import io
import math
import re
from pathlib import Path
from typing import Iterable

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    import matplotlib.pyplot as plt
except Exception as exc:  # pragma: no cover - environment-dependent import failure
    plt = None
    MATPLOTLIB_IMPORT_ERROR = exc
else:
    MATPLOTLIB_IMPORT_ERROR = None

if plt is not None:
    from openpyxl.drawing.image import Image as XLImage
else:
    XLImage = None


SULCUS_CLASSES = ("Primary", "Secondary", "Tertiary", "Unclassified")
CORE_METRICS = ("area", "perimeter", "LGI", "Compactness")
COUNT_METRICS = ("Sulci_count", "Primary_count", "Secondary_count", "Tertiary_count", "Unclassified_count")
DEPTH_METRIC_PREFIXES = ("min_depth_", "max_depth_", "total_depth_", "mean_depth_")
METADATA_LABELS = {
    "PixelSize:",
    "PixelSizeUnits:",
    "KernelSize:",
    "ScalebarMeasuredPixels:",
    "ScalebarRealWorldLength:",
    "ScalebarRealWorldUnit:",
    "PixelSizeUnits",
    "KernelSize",
}
WORKBOOK_PATTERN = re.compile(r"week(?P<week>\d+)_(?P<axis>[A-Za-z]+)_Batch_Allmarks\.xlsx$", re.IGNORECASE)
ANALYSIS_SHEET_NAME = "Analysis"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Analyze each master measurement report workbook independently and "
            "write summary stats plus boxplots back into the same Excel file."
        )
    )
    parser.add_argument(
        "--input-root",
        type=Path,
        default=Path("measurements"),
        help="Root folder containing week/axis measurement report workbooks.",
    )
    parser.add_argument(
        "--axes",
        nargs="+",
        default=["axial", "coronal", "sagittal"],
        help="Axis names to include. Defaults to axial, coronal, and sagittal.",
    )
    parser.add_argument(
        "--weeks",
        nargs="+",
        type=int,
        help="Optional weeks to include, for example: --weeks 24 25 26",
    )
    parser.add_argument(
        "--report-glob",
        default="week*_Batch_Allmarks.xlsx",
        help="Glob used below input-root to discover workbooks.",
    )
    parser.add_argument(
        "--sheet-name",
        default=ANALYSIS_SHEET_NAME,
        help="Worksheet name used for the generated analysis output.",
    )
    return parser.parse_args()


def discover_workbooks(input_root: Path, report_glob: str, axes: set[str], weeks: set[int] | None) -> list[Path]:
    paths: list[Path] = []
    for path in sorted(input_root.rglob(report_glob)):
        match = WORKBOOK_PATTERN.match(path.name)
        if not match:
            continue
        axis = match.group("axis").lower()
        week = int(match.group("week"))
        if axis not in axes:
            continue
        if weeks is not None and week not in weeks:
            continue
        paths.append(path)
    return paths


def is_data_row(file_value: object) -> bool:
    if not isinstance(file_value, str):
        return False
    text = file_value.strip()
    if not text:
        return False
    if text in METADATA_LABELS:
        return False
    if text.endswith(":"):
        return False
    return True


def clean_dataframe(df: pd.DataFrame, workbook_path: Path) -> pd.DataFrame:
    if "File" not in df.columns:
        raise ValueError(f"{workbook_path} does not contain a 'File' column.")

    keep_cols = []
    for col in df.columns:
        if col in {"Metric", "Mean"}:
            continue
        if col is None:
            continue
        if isinstance(col, str) and col.startswith("Unnamed:"):
            continue
        keep_cols.append(col)

    df = df[keep_cols].copy()
    df = df[df["File"].map(is_data_row)].copy()
    return df.reset_index(drop=True)


def parse_workbook(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path, engine="openpyxl")
    df = clean_dataframe(df, path)

    match = WORKBOOK_PATTERN.match(path.name)
    if match is None:
        raise ValueError(f"Workbook name does not match expected pattern: {path.name}")

    df["week"] = int(match.group("week"))
    df["axis"] = match.group("axis").lower()
    df["workbook"] = path.name
    return df


def infer_depth_unit(df: pd.DataFrame) -> str:
    for col in df.columns:
        if not isinstance(col, str):
            continue
        match = re.search(r"_([A-Za-z]+)$", col)
        if match and col.startswith((*DEPTH_METRIC_PREFIXES, "Primary_v1_", "Secondary_v1_")):
            return match.group(1)
    return "mm"


def depth_metric_columns(df: pd.DataFrame, unit: str) -> list[str]:
    return [f"{prefix}{unit}" for prefix in DEPTH_METRIC_PREFIXES if f"{prefix}{unit}" in df.columns]


def round_count_series(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce").round().astype("Int64")


def existing_columns(df: pd.DataFrame, names: Iterable[str]) -> list[str]:
    return [name for name in names if name in df.columns]


def expand_sulcus_values(row: pd.Series, prefix: str, unit: str) -> list[float]:
    count_col = f"{prefix}_count"
    count_raw = row.get(count_col)
    if pd.isna(count_raw):
        return []

    count = int(round(float(count_raw)))
    if count <= 0:
        return []

    raw_cols = [f"{prefix}_v1_{unit}", f"{prefix}_v2_{unit}", f"{prefix}_v3_{unit}"]
    raw_values = [
        float(row[col])
        for col in raw_cols
        if col in row.index and pd.notna(row[col])
    ]
    if raw_values:
        return raw_values[:count]

    min_col = f"{prefix}_min_{unit}"
    max_col = f"{prefix}_max_{unit}"
    mean_col = f"{prefix}_mean_{unit}"
    min_value = row.get(min_col)
    max_value = row.get(max_col)
    mean_value = row.get(mean_col)

    if pd.notna(min_value) and pd.notna(max_value) and pd.notna(mean_value):
        min_float = float(min_value)
        max_float = float(max_value)
        mean_float = float(mean_value)
        if count == 1:
            return [mean_float]
        if count == 2:
            return [min_float, max_float]
        filler_count = count - 2
        filler = ((mean_float * count) - min_float - max_float) / filler_count
        return [min_float] + [filler] * filler_count + [max_float]

    if pd.notna(mean_value):
        return [float(mean_value)] * count
    if pd.notna(min_value) and pd.notna(max_value):
        if count == 1:
            return [float(min_value)]
        return [float(min_value)] + [float(max_value)] * (count - 1)
    return []


def build_sulcus_value_table(df: pd.DataFrame, unit: str) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for _, row in df.iterrows():
        for prefix in SULCUS_CLASSES:
            for value in expand_sulcus_values(row, prefix, unit):
                rows.append(
                    {
                        "slice_file": row["File"],
                        "sulcus_class": prefix.lower(),
                        "value": float(value),
                    }
                )
    # Always return the expected schema, even when empty (e.g. cropped slices,
    # which are all "unclassified" and carry no per-class sulcus columns), so
    # downstream lookups on "sulcus_class" never raise KeyError.
    return pd.DataFrame(rows, columns=["slice_file", "sulcus_class", "value"])


def metric_summary(df: pd.DataFrame, metrics: Iterable[str]) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for metric in metrics:
        if metric not in df.columns:
            continue
        values = pd.to_numeric(df[metric], errors="coerce").dropna()
        rows.append(
            {
                "metric": metric,
                "n": int(values.shape[0]),
                "mean": float(values.mean()) if not values.empty else math.nan,
                "std": float(values.std(ddof=1)) if len(values) > 1 else math.nan,
                "min": float(values.min()) if not values.empty else math.nan,
                "max": float(values.max()) if not values.empty else math.nan,
            }
        )
    return pd.DataFrame(rows)


def rounded_counts_table(df: pd.DataFrame) -> pd.DataFrame:
    cols = ["File"] + existing_columns(df, COUNT_METRICS)
    rounded = df[cols].copy()
    for col in existing_columns(df, COUNT_METRICS):
        rounded[f"{col}_rounded"] = round_count_series(df[col])
    return rounded


def count_summary(rounded_df: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for col in rounded_df.columns:
        if not col.endswith("_rounded"):
            continue
        values = pd.to_numeric(rounded_df[col], errors="coerce").dropna().astype(int)
        rows.append(
            {
                "metric": col,
                "n": int(values.shape[0]),
                "mean": float(values.mean()) if not values.empty else math.nan,
                "std": float(values.std(ddof=1)) if len(values) > 1 else math.nan,
                "min": int(values.min()) if not values.empty else pd.NA,
                "max": int(values.max()) if not values.empty else pd.NA,
            }
        )
    return pd.DataFrame(rows)


def depth_summary(df: pd.DataFrame, unit: str) -> pd.DataFrame:
    return metric_summary(df, depth_metric_columns(df, unit))


def per_class_sulcus_summary(sulcus_values_df: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for class_name in [name.lower() for name in SULCUS_CLASSES]:
        values = pd.to_numeric(
            sulcus_values_df.loc[sulcus_values_df["sulcus_class"] == class_name, "value"],
            errors="coerce",
        ).dropna()
        rows.append(
            {
                "metric": f"{class_name}_sulcus_values",
                "n": int(values.shape[0]),
                "mean": float(values.mean()) if not values.empty else math.nan,
                "std": float(values.std(ddof=1)) if len(values) > 1 else math.nan,
                "min": float(values.min()) if not values.empty else math.nan,
                "max": float(values.max()) if not values.empty else math.nan,
            }
        )
    return pd.DataFrame(rows)


BOX_COLOR = "#9ecae1"
SCATTER_COLOR = "#2E6DA4"


def _save_figure(fig) -> io.BytesIO:
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=180)
    plt.close(fig)
    buf.seek(0)
    return buf


def _annotate_median(ax, position: float, median: float, fmt: str = "{:.3g}") -> None:
    ax.text(
        position,
        median,
        fmt.format(median),
        ha="center",
        va="bottom",
        fontsize=8,
        fontweight="bold",
        color="#1A3A5C",
    )


def plot_boxplot(values: pd.Series, title: str, ylabel: str) -> io.BytesIO | None:
    if plt is None:
        return None
    clean = pd.to_numeric(values, errors="coerce").dropna()
    if clean.empty:
        return None

    fig, ax = plt.subplots(figsize=(7, 4))
    bp = ax.boxplot(clean, patch_artist=True, boxprops={"facecolor": BOX_COLOR})
    ax.set_xticklabels([])
    ax.tick_params(axis="x", length=0)
    _annotate_median(ax, 1, float(bp["medians"][0].get_ydata()[0]))
    ax.set_title(title)
    ax.set_ylabel(ylabel)
    ax.grid(axis="y", linestyle=":", alpha=0.5)
    fig.tight_layout()
    return _save_figure(fig)


def plot_grouped_sulcus_boxplot(sulcus_values_df: pd.DataFrame, ylabel: str) -> io.BytesIO | None:
    if plt is None:
        return None
    grouped = []
    labels = []
    for class_name in [name.lower() for name in SULCUS_CLASSES]:
        values = pd.to_numeric(
            sulcus_values_df.loc[sulcus_values_df["sulcus_class"] == class_name, "value"],
            errors="coerce",
        ).dropna()
        if values.empty:
            continue
        grouped.append(values)
        labels.append(class_name)

    if not grouped:
        return None

    fig, ax = plt.subplots(figsize=(8.5, 4.5))
    bp = ax.boxplot(grouped, labels=labels, patch_artist=True,
                    boxprops={"facecolor": BOX_COLOR})
    for i, median_line in enumerate(bp["medians"]):
        _annotate_median(ax, i + 1, float(median_line.get_ydata()[0]))
    ax.set_title("Sulcus value distributions by class")
    ax.set_ylabel(ylabel)
    ax.grid(axis="y", linestyle=":", alpha=0.5)
    fig.tight_layout()
    return _save_figure(fig)


def plot_grouped_count_boxplot(rounded_df: pd.DataFrame) -> io.BytesIO | None:
    if plt is None:
        return None
    grouped = []
    labels = []
    count_columns = (
        ("Primary_count_rounded", "primary"),
        ("Secondary_count_rounded", "secondary"),
        ("Tertiary_count_rounded", "tertiary"),
        ("Unclassified_count_rounded", "unclassified"),
    )
    for column_name, label in count_columns:
        if column_name not in rounded_df.columns:
            continue
        values = pd.to_numeric(rounded_df[column_name], errors="coerce").dropna()
        if values.empty:
            continue
        grouped.append(values)
        labels.append(label)

    if not grouped:
        return None

    fig, ax = plt.subplots(figsize=(9.5, 5.0))
    positions = list(range(1, len(grouped) + 1))
    bp = ax.boxplot(grouped, labels=labels, positions=positions, patch_artist=True,
                    boxprops={"facecolor": BOX_COLOR})
    for i, (pos, values) in enumerate(zip(positions, grouped)):
        offsets = np.linspace(-0.08, 0.08, len(values)) if len(values) > 1 else np.array([0.0])
        ax.scatter(
            pos + offsets,
            values,
            s=20,
            color=SCATTER_COLOR,
            alpha=0.7,
            edgecolors="none",
            zorder=3,
        )
        _annotate_median(ax, pos, float(bp["medians"][i].get_ydata()[0]), fmt="{:.0f}")
    ax.set_title("Sulcus count distributions by fold class")
    ax.set_ylabel("count per slice")
    max_count = max(int(values.max()) for values in grouped)
    ax.set_ylim(bottom=-0.1, top=max_count + 0.35)
    ax.set_yticks(list(range(0, max_count + 1)))
    ax.grid(axis="y", linestyle=":", alpha=0.5)
    fig.tight_layout()
    return _save_figure(fig)


HEADER_FILL = PatternFill("solid", fgColor="2E6DA4")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
SECTION_FILL = PatternFill("solid", fgColor="D9E8F5")
SECTION_FONT = Font(bold=True, color="1A3A5C", size=11)
ROW_FILL_ODD = PatternFill("solid", fgColor="FFFFFF")
ROW_FILL_EVEN = PatternFill("solid", fgColor="EEF4FB")
BORDER_SIDE = Side(style="thin", color="B0C4D8")
CELL_BORDER = Border(
    left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE
)


def _style_cell(cell, fill=None, font=None, border=True, align_center=False) -> None:
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if border:
        cell.border = CELL_BORDER
    if align_center:
        cell.alignment = Alignment(horizontal="center")


def _autofit_columns(ws) -> None:
    col_widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            text_len = len(str(cell.value))
            col_widths[cell.column] = max(col_widths.get(cell.column, 0), text_len)
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width + 3, 50)


def _set_wrapped_row_height(ws, row: int, columns: Iterable[int], base_height: int = 18) -> None:
    max_lines = 1
    for col in columns:
        cell = ws.cell(row=row, column=col)
        if cell.value is None:
            continue
        text = str(cell.value)
        width = ws.column_dimensions[get_column_letter(col)].width or 10
        approx_chars = max(12, int(width * 1.15))
        wrapped_lines = sum(max(1, math.ceil(len(part) / approx_chars)) for part in text.splitlines() or [""])
        max_lines = max(max_lines, wrapped_lines)
    ws.row_dimensions[row].height = base_height * max_lines


def write_dataframe(ws, start_row: int, start_col: int, title: str, df: pd.DataFrame) -> tuple[int, int]:
    """Returns (next_free_row, rightmost_column_used)."""
    num_cols = max(len(df.columns), 1)
    end_col = start_col + num_cols - 1

    title_cell = ws.cell(row=start_row, column=start_col, value=title)
    _style_cell(title_cell, fill=SECTION_FILL, font=SECTION_FONT, border=False)
    if num_cols > 1:
        ws.merge_cells(
            start_row=start_row, start_column=start_col,
            end_row=start_row, end_column=end_col,
        )

    if df.empty:
        ws.cell(row=start_row + 1, column=start_col, value="No data")
        return start_row + 3, end_col

    header_row = start_row + 1
    for offset, col_name in enumerate(df.columns):
        cell = ws.cell(row=header_row, column=start_col + offset, value=str(col_name))
        _style_cell(cell, fill=HEADER_FILL, font=HEADER_FONT)

    for row_offset, row in enumerate(df.itertuples(index=False), start=1):
        fill = ROW_FILL_EVEN if row_offset % 2 == 0 else ROW_FILL_ODD
        for col_offset, value in enumerate(row):
            cell = ws.cell(
                row=header_row + row_offset,
                column=start_col + col_offset,
                value=normalize_excel_value(value),
            )
            _style_cell(cell, fill=fill)

    return header_row + len(df) + 2, end_col


def normalize_excel_value(value: object) -> object:
    if pd.isna(value):
        return None
    return value


def replace_analysis_sheet(workbook_path: Path, sheet_name: str):
    wb = load_workbook(workbook_path)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)
    return wb, ws


def add_image(ws, image_bytes: io.BytesIO | None, anchor: str, width: int = 520, height: int = 360) -> None:
    if image_bytes is None or XLImage is None:
        return
    img = XLImage(image_bytes)
    img.width = width
    img.height = height
    ws.add_image(img, anchor)


def analyze_workbook(path: Path, sheet_name: str) -> None:
    df = parse_workbook(path)
    depth_unit = infer_depth_unit(df)

    for col in existing_columns(df, CORE_METRICS):
        df[col] = pd.to_numeric(df[col], errors="coerce")

    rounded_counts_df = rounded_counts_table(df)
    core_summary_df = metric_summary(df, CORE_METRICS)
    count_summary_df = count_summary(rounded_counts_df)
    depth_summary_df = depth_summary(df, depth_unit)
    sulcus_values_df = build_sulcus_value_table(df, depth_unit)
    has_sulcus_values = not sulcus_values_df.empty
    sulcus_summary_df = per_class_sulcus_summary(sulcus_values_df) if has_sulcus_values else None

    wb, ws = replace_analysis_sheet(path, sheet_name)

    match = WORKBOOK_PATTERN.match(path.name)
    assert match is not None
    week = int(match.group("week"))
    axis = match.group("axis").lower()

    # ── Metadata info card ───────────────────────────────────────────────────
    CARD_TITLE_FILL = PatternFill("solid", fgColor="1A3A5C")
    CARD_TITLE_FONT = Font(bold=True, color="FFFFFF", size=12)
    CARD_LABEL_FILL = PatternFill("solid", fgColor="D9E8F5")
    CARD_LABEL_FONT = Font(bold=True, color="1A3A5C", size=10)
    CARD_VALUE_FILL = PatternFill("solid", fgColor="EEF4FB")
    CARD_VALUE_FONT = Font(color="1A1A1A", size=10)

    title_cell = ws["A1"]
    title_cell.value = "Workbook Analysis"
    title_cell.fill = CARD_TITLE_FILL
    title_cell.font = CARD_TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 22

    card_rows = [
        ("Workbook", path.name),
        ("Week", week),
        ("Axis", axis),
        ("Slice rows analyzed", len(df)),
    ]
    note_parts = [
        "Per-class sulcus plots may use values reconstructed from count + min/max/mean."
    ]
    if not has_sulcus_values:
        note_parts.append(
            "Per-class sulcus tables/plots were skipped because this cropped-slice workbook has no per-class sulcus columns."
        )
    if MATPLOTLIB_IMPORT_ERROR is not None:
        note_parts.append(
            "Embedded plot images were skipped because matplotlib was unavailable in this environment."
        )
    note_text = " ".join(note_parts)

    for idx, (label, value) in enumerate(card_rows, start=2):
        lbl = ws.cell(row=idx, column=1, value=label)
        lbl.fill = CARD_LABEL_FILL
        lbl.font = CARD_LABEL_FONT
        lbl.border = CELL_BORDER
        val = ws.cell(row=idx, column=2, value=value)
        val.fill = CARD_VALUE_FILL
        val.font = CARD_VALUE_FONT
        val.border = CELL_BORDER

    note_label = ws.cell(row=6, column=1, value="Note")
    note_label.fill = CARD_LABEL_FILL
    note_label.font = CARD_LABEL_FONT
    note_label.border = CELL_BORDER
    note_value = ws.cell(row=6, column=2, value=note_text)
    note_value.fill = CARD_VALUE_FILL
    note_value.font = CARD_VALUE_FONT
    note_value.border = CELL_BORDER
    note_value.alignment = Alignment(wrap_text=True)

    # ── Data tables ──────────────────────────────────────────────────────────
    next_row = 8
    max_data_col = 1
    next_row, used_col = write_dataframe(ws, next_row, 1, "Core Metric Summary", core_summary_df)
    max_data_col = max(max_data_col, used_col)
    next_row, used_col = write_dataframe(ws, next_row, 1, "Rounded Sulcus Counts Per Slice", rounded_counts_df)
    max_data_col = max(max_data_col, used_col)
    next_row, used_col = write_dataframe(ws, next_row, 1, "Rounded Sulcus Count Summary", count_summary_df)
    max_data_col = max(max_data_col, used_col)
    next_row, used_col = write_dataframe(ws, next_row, 1, "Depth Metric Summary", depth_summary_df)
    max_data_col = max(max_data_col, used_col)
    if has_sulcus_values:
        next_row, used_col = write_dataframe(ws, next_row, 1, "Sulcus Value Summary", sulcus_summary_df)
        max_data_col = max(max_data_col, used_col)

    # ── Plot anchors: placed 2 columns right of the widest table ─────────────
    # Each plot group is ~9 columns wide at default column width (~64 px each).
    plot_col_1 = max_data_col + 2          # core metric plots
    plot_col_2 = plot_col_1 + 9           # per-class sulcus value plots
    plot_col_3 = plot_col_2 + 9           # grouped summary plots

    def _col_anchor(col: int, row: int) -> str:
        return f"{get_column_letter(col)}{row}"

    PLOT_ROW_STEP = 22
    core_metric_plots = ["area", "perimeter", "LGI", "Compactness"]
    for i, metric in enumerate(core_metric_plots):
        if metric not in df.columns:
            continue
        add_image(
            ws,
            plot_boxplot(df[metric], f"{metric} distribution", metric),
            _col_anchor(plot_col_1, 2 + i * PLOT_ROW_STEP),
        )

    for i, metric in enumerate(depth_metric_columns(df, depth_unit)):
        add_image(
            ws,
            plot_boxplot(df[metric], f"{metric} distribution", metric),
            _col_anchor(plot_col_1, 2 + (len(core_metric_plots) + i) * PLOT_ROW_STEP),
        )

    if has_sulcus_values:
        sulcus_classes_ordered = ["primary", "secondary", "tertiary", "unclassified"]
        for i, class_name in enumerate(sulcus_classes_ordered):
            class_values = sulcus_values_df.loc[sulcus_values_df["sulcus_class"] == class_name, "value"]
            add_image(
                ws,
                plot_boxplot(
                    class_values,
                    f"{class_name.capitalize()} sulcus values distribution",
                    f"depth ({depth_unit})",
                ),
                _col_anchor(plot_col_2, 2 + i * PLOT_ROW_STEP),
            )

        add_image(
            ws,
            plot_grouped_sulcus_boxplot(sulcus_values_df, f"depth ({depth_unit})"),
            _col_anchor(plot_col_3, 2),
            width=620,
            height=360,
        )
    add_image(
        ws,
        plot_grouped_count_boxplot(rounded_counts_df),
        _col_anchor(plot_col_3, 2 + PLOT_ROW_STEP),
        width=680,
        height=400,
    )

    _autofit_columns(ws)
    _set_wrapped_row_height(ws, 6, [1, 2])
    wb.save(path)


def main() -> int:
    args = parse_args()
    input_root = args.input_root
    axes = {axis.strip().lower() for axis in args.axes}
    weeks = set(args.weeks) if args.weeks else None

    report_paths = discover_workbooks(input_root, args.report_glob, axes, weeks)
    if not report_paths:
        raise SystemExit(f"No matching workbooks found under {input_root}")

    for path in report_paths:
        analyze_workbook(path, args.sheet_name)
        print(f"Updated workbook: {path}")

    if MATPLOTLIB_IMPORT_ERROR is not None:
        print(
            "Warning: matplotlib could not be imported, so analysis sheets were written "
            "without embedded plot images."
        )
        print(f"Import error: {MATPLOTLIB_IMPORT_ERROR}")

    print(f"Analyzed {len(report_paths)} workbooks.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
