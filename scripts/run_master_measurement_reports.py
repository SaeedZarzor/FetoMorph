from __future__ import annotations

import argparse
import json
import logging
import shutil
import sys
from pathlib import Path
from typing import Any

import cv2
import numpy as np
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from functions.measurement_batch import process_on_images_batch
from helpers.helpers import (
    compute_kernel_convex,
    compactness_2D,
    image_annotation_style,
    SULCUS_CLASS_COLORS,
)
from constants import BINARY_THRESHOLD_DEFAULT, DEFECT_FIXED_POINT


IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff"}
DEFAULT_AXES = ("axial", "coronal", "sagittal")
DEFAULT_WEEKS = tuple(range(24, 39))
DEFAULT_CONFIG = {
    "input_root": str(REPO_ROOT / "Examples" / "full_slices"),
    "output_root": str(REPO_ROOT / "measurements"),
    "section_label": "Filled_2D_sections",
    "weeks": list(DEFAULT_WEEKS),
    "axes": list(DEFAULT_AXES),
    "pixel_size": 20.0 / 42.0,
    "scalebar_measured_pixels": 42.0,
    "scalebar_real_world_length": 20.0,
    "kernel_size": 25,
    "cnt_threshold": 2000,
    "unit": "mm",
    "auto_scalebar": False,
    "single_pass_metrics": False,
    "review": False,
    "review_upscale_min": 600,
    "area_close_mm": 0.0,
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Run the master-branch FetoMorph batch measurement pipeline across "
            "week/axis folders and write one report workbook per folder."
        )
    )
    parser.add_argument("--config", type=Path, help="Optional JSON config file.")
    parser.add_argument("--input-root", type=Path, help="Root folder containing week/axis image folders.")
    parser.add_argument("--output-root", type=Path, help="Root folder where reports will be written.")
    parser.add_argument(
        "--section-label",
        type=str,
        help='Section folder placed under each week, for example "Filled_2D_sections".',
    )
    parser.add_argument("--weeks", nargs="+", type=int, help="Weeks to process, for example: --weeks 24 25")
    parser.add_argument(
        "--axes",
        nargs="+",
        help="Axes to process, for example: --axes axial coronal sagittal",
    )
    parser.add_argument("--pixel-size", type=float, help="Physical pixel size in unit/pixel.")
    parser.add_argument(
        "--scalebar-measured-pixels",
        type=float,
        help="Measured scalebar length in pixels (for scale-from-scalebar calibration).",
    )
    parser.add_argument(
        "--scalebar-real-world-length",
        type=float,
        help="Real-world scalebar length in current unit (for scale-from-scalebar calibration).",
    )
    parser.add_argument("--kernel-size", type=int, help="Morphology kernel size.")
    parser.add_argument("--cnt-threshold", type=float, help="Minimum contour area threshold in pixels.")
    parser.add_argument("--unit", type=str, help='Length unit label, for example "mm".')
    parser.add_argument(
        "--auto-scalebar",
        action="store_true",
        default=None,
        help="Derive pixel_size per folder from each image's embedded scalebar bar "
             "(falls back to the configured pixel_size when no bar is detected).",
    )
    parser.add_argument(
        "--single-pass-metrics",
        action="store_true",
        default=None,
        help="After the batch run, recompute area/perimeter/perimeter_convex/LGI/"
             "Compactness with a single pass (no LGI threshold-climbing retry), so "
             "small cropped brains are never erased. Overwrites those columns.",
    )
    parser.add_argument(
        "--review",
        action="store_true",
        default=None,
        help="Also write enlarged annotated review PNGs (corrected segmentation, "
             "scalebar kept) to a 'review' sub-folder beside each report, for "
             "visual inspection before trusting the numbers.",
    )
    parser.add_argument(
        "--review-upscale-min",
        type=int,
        help="Target minimum side length (px) for review PNGs. Default 600.",
    )
    parser.add_argument(
        "--area-close-mm",
        type=float,
        help="Option B: close the white cortex ring into the area. The reported "
             "area is measured on the brain after a morphological close of this "
             "physical distance (mm), scaled per folder via pixel_size. LGI and "
             "perimeter stay on the folded boundary. 0 disables (area = tissue only).",
    )
    parser.add_argument(
        "--log-level",
        default="INFO",
        choices=("DEBUG", "INFO", "WARNING", "ERROR"),
        help="Console logging verbosity.",
    )
    return parser.parse_args()


def load_config(path: Path | None) -> dict[str, Any]:
    if path is None:
        return {}
    with path.open("r", encoding="utf-8") as fh:
        data = json.load(fh)
    if not isinstance(data, dict):
        raise ValueError("Config file must contain a JSON object.")
    return data


def resolve_settings(args: argparse.Namespace) -> dict[str, Any]:
    cfg = DEFAULT_CONFIG.copy()
    cfg.update(load_config(args.config))

    overrides = {
        "input_root": args.input_root,
        "output_root": args.output_root,
        "section_label": args.section_label,
        "weeks": args.weeks,
        "axes": args.axes,
        "pixel_size": args.pixel_size,
        "scalebar_measured_pixels": args.scalebar_measured_pixels,
        "scalebar_real_world_length": args.scalebar_real_world_length,
        "kernel_size": args.kernel_size,
        "cnt_threshold": args.cnt_threshold,
        "unit": args.unit,
        "auto_scalebar": args.auto_scalebar,
        "single_pass_metrics": args.single_pass_metrics,
        "review": args.review,
        "review_upscale_min": args.review_upscale_min,
        "area_close_mm": args.area_close_mm,
    }
    for key, value in overrides.items():
        if value is not None:
            cfg[key] = value

    cfg["input_root"] = Path(cfg["input_root"]).resolve()
    cfg["output_root"] = Path(cfg["output_root"]).resolve()
    cfg["section_label"] = str(cfg["section_label"]).strip()
    cfg["weeks"] = [int(w) for w in cfg["weeks"]]
    cfg["axes"] = [str(a).strip().lower() for a in cfg["axes"]]
    scalebar_px = cfg.get("scalebar_measured_pixels")
    scalebar_real = cfg.get("scalebar_real_world_length")
    if scalebar_px is not None or scalebar_real is not None:
        if scalebar_px is None or scalebar_real is None:
            raise ValueError(
                "Both scalebar_measured_pixels and scalebar_real_world_length must be provided together."
            )
        scalebar_px = float(scalebar_px)
        scalebar_real = float(scalebar_real)
        if scalebar_px <= 0 or scalebar_real <= 0:
            raise ValueError("Scalebar calibration values must be positive numbers.")
        cfg["scalebar_measured_pixels"] = scalebar_px
        cfg["scalebar_real_world_length"] = scalebar_real
        cfg["pixel_size"] = scalebar_real / scalebar_px
    else:
        cfg["scalebar_measured_pixels"] = None
        cfg["scalebar_real_world_length"] = None
        cfg["pixel_size"] = float(cfg["pixel_size"])

    cfg["kernel_size"] = int(cfg["kernel_size"])
    cfg["cnt_threshold"] = float(cfg["cnt_threshold"])
    cfg["unit"] = str(cfg["unit"]).strip()
    cfg["auto_scalebar"] = bool(cfg.get("auto_scalebar", False))
    cfg["single_pass_metrics"] = bool(cfg.get("single_pass_metrics", False))
    cfg["review"] = bool(cfg.get("review", False))
    cfg["review_upscale_min"] = int(cfg.get("review_upscale_min", 600))
    cfg["area_close_mm"] = float(cfg.get("area_close_mm", 0.0) or 0.0)
    return cfg


def find_image_files(folder: Path) -> list[Path]:
    return sorted(
        [p for p in folder.iterdir() if p.is_file() and p.suffix.lower() in IMAGE_EXTS],
        key=lambda p: p.name.lower(),
    )


# ---------------------------------------------------------------------------
# Foreground segmentation (scripts-local).
#
# The shared functions/measurement_batch binarisation keeps a pixel as "brain"
# only when its grayscale value is dark enough (THRESH_BINARY_INV @ 200). That
# is correct for a grayscale render but wrong for these coloured label-map
# crops: bright labels such as cyan (gray ~212) fall just over the threshold
# and are dropped, and the inner white cortex (gray 255) is dropped entirely.
# The recompute path below replaces that rule WITHOUT editing the shared module.
# ---------------------------------------------------------------------------
WHITE_BG_CHANNEL_MIN = 235  # a pixel is "background white" when every BGR channel >= this


def _border_is_white(img: np.ndarray, border: int = 3, white_min: int = WHITE_BG_CHANNEL_MIN) -> bool:
    """True when the image border is predominantly near-white.

    Distinguishes white-background colour label maps (the brain crops, where the
    region of interest is *any* non-white pixel) from dark-background grayscale
    renders (where the legacy brightness threshold is the correct rule).
    """
    edges = np.concatenate([
        img[:border, :, :].reshape(-1, 3),
        img[-border:, :, :].reshape(-1, 3),
        img[:, :border, :].reshape(-1, 3),
        img[:, -border:, :].reshape(-1, 3),
    ])
    return bool(np.all(edges >= white_min, axis=1).mean() > 0.5)


def segment_foreground(img: np.ndarray, white_min: int = WHITE_BG_CHANNEL_MIN) -> np.ndarray:
    """Binary brain mask (uint8, 255 = brain) for one slice image.

    White-background colour label maps: foreground is every non-white pixel — so
    bright labels such as cyan are kept rather than dropped by a brightness cut —
    PLUS any white region fully enclosed by tissue, which is the inner cortex and
    gets filled in. White that opens outward to the image border stays background,
    so the sulcal concavities (and therefore the LGI / sulcus-depth signal) are
    preserved: this is the "fill enclosed only, keep folds" behaviour.

    Dark-background grayscale renders fall back to the original
    THRESH_BINARY_INV @ BINARY_THRESHOLD_DEFAULT rule, leaving full-slice
    behaviour unchanged.
    """
    if not _border_is_white(img, white_min=white_min):
        gray = cv2.cvtColor(img, cv2.COLOR_RGB2GRAY)
        _, bw = cv2.threshold(gray, BINARY_THRESHOLD_DEFAULT, 255, 1)
        return bw

    nonwhite = np.where(np.all(img >= white_min, axis=2), 0, 255).astype(np.uint8)
    # Fill white fully enclosed by tissue (interior cortex); white that reaches
    # the border (exterior background + open sulcal valleys) is left alone.
    return _fill_interior_holes(nonwhite)


def _fill_interior_holes(mask: np.ndarray) -> np.ndarray:
    """Flood the foreground over every background component that does not touch
    the image border (i.e. holes fully enclosed by foreground)."""
    num, labels = cv2.connectedComponents(cv2.bitwise_not(mask))
    border_labels = set(
        np.concatenate([labels[0, :], labels[-1, :], labels[:, 0], labels[:, -1]]).tolist()
    )
    filled = mask.copy()
    for lbl in range(1, num):
        if lbl not in border_labels:
            filled[labels == lbl] = 255
    return filled


def area_close_kernel_px(area_close_mm: float, pixel_size: float, min_px: int) -> int:
    """Close-kernel size (px) for the area mask, from a physical closing distance.

    Scaling by ``pixel_size`` keeps the cortex-ring fill consistent across the
    different per-folder zooms; never smaller than ``min_px`` (the LGI kernel).
    """
    if area_close_mm <= 0 or pixel_size <= 0:
        return min_px
    return max(min_px, int(round(area_close_mm / pixel_size)))


def closed_brain_mask(kept_contours, shape, kernel_px: int) -> np.ndarray:
    """Solid brain mask: fill the kept contours, morph-close to seal the cortex
    ring / sulcal mouths, then fill any newly enclosed holes."""
    inner = np.zeros(shape, np.uint8)
    cv2.drawContours(inner, kept_contours, -1, 255, thickness=cv2.FILLED)
    closed = cv2.morphologyEx(inner, cv2.MORPH_CLOSE, compute_kernel_convex(kernel_px))
    return _fill_interior_holes(closed)


def render_review_image(
    img: np.ndarray,
    kernel_size: int,
    cnt_threshold: float,
    pixel_size: float,
    unit: str = "mm",
    area_close_mm: float = 0.0,
    *,
    upscale_min: int = 600,
) -> np.ndarray:
    """Enlarged annotated review image, drawn like the full-slice report.

    Reproduces the four annotations of ``measurement_batch`` — red inner brain
    contour, green envelope, blue convexity-defect chords (the outer line that
    wraps the whole region), and sulci-depth markers at each kept defect's far
    point — but on the corrected non-white segmentation, so cyan and the inner
    cortex are no longer dropped.

    Under option B (``area_close_mm > 0``) the green outline is the
    cortex-ring-filled brain that ``area`` is measured on; otherwise it is the
    light LGI envelope. Red always shows the folded boundary.

    The whole frame (scalebar included) is scaled up with nearest-neighbour for
    legibility; metrics are computed elsewhere at native resolution, so
    calibration is unaffected.
    """
    mask = segment_foreground(img)
    contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    kept = [c for c in contours if cv2.contourArea(c) > cnt_threshold]

    if area_close_mm > 0:
        env_mask = closed_brain_mask(
            kept, mask.shape, area_close_kernel_px(area_close_mm, pixel_size, kernel_size)
        )
    else:
        inner = np.zeros(mask.shape, np.uint8)
        cv2.drawContours(inner, kept, -1, 255, thickness=cv2.FILLED)
        env_mask = cv2.morphologyEx(inner, cv2.MORPH_CLOSE, compute_kernel_convex(kernel_size))
    conv, _ = cv2.findContours(env_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    kept_conv = [c for c in conv if cv2.contourArea(c) > cnt_threshold]

    annotated = img.copy()
    h, w = annotated.shape[:2]
    thickness, _font_scale, radius_px = image_annotation_style(w, h, style="bold")

    cv2.drawContours(annotated, kept_conv, -1, (0, 255, 0), thickness)  # green envelope
    cv2.drawContours(annotated, kept, -1, (0, 0, 255), thickness)       # red inner contour

    # Blue convexity-defect chords + sulci markers (cropped sub-slices are not
    # full MRI slices, so the fixed-depth keep rule and "unclassified" marker
    # are used, matching measurement_batch's non-percent branch).
    min_keep = 0.5 if unit == "mm" else 0.05 if unit == "cm" else 0.0
    marker_color = SULCUS_CLASS_COLORS["unclassified"]
    for cnt in kept:
        hull = cv2.convexHull(cnt, returnPoints=False, clockwise=True)
        if hull is None or len(hull) < 3 or len(cnt) <= 3 or not np.all(np.diff(hull.ravel()) > 0):
            continue
        defects = cv2.convexityDefects(cnt, hull)
        if defects is None:
            continue
        for i in range(defects.shape[0]):
            s, e, f, d = defects[i, 0]
            start, end, far = tuple(cnt[s][0]), tuple(cnt[e][0]), tuple(cnt[f][0])
            cv2.line(annotated, start, end, [255, 0, 0], thickness)  # blue outer chord
            if (d * pixel_size / DEFECT_FIXED_POINT) > min_keep:
                cv2.circle(annotated, far, radius_px, marker_color, -1)

    factor = max(1, int(round(upscale_min / max(min(h, w), 1))))
    return cv2.resize(annotated, (w * factor, h * factor), interpolation=cv2.INTER_NEAREST)


def render_review_images(
    image_dir: Path,
    output_dir: Path,
    kernel_size: int,
    cnt_threshold: float,
    pixel_size: float,
    unit: str = "mm",
    area_close_mm: float = 0.0,
    *,
    upscale_min: int = 600,
) -> int:
    """Write one enlarged review PNG per image into ``<output_dir>/review``."""
    review_dir = output_dir / "review"
    review_dir.mkdir(parents=True, exist_ok=True)
    written = 0
    for path in find_image_files(image_dir):
        img = cv2.imread(str(path))
        if img is None:
            continue
        big = render_review_image(
            img, kernel_size, cnt_threshold, pixel_size, unit, area_close_mm,
            upscale_min=upscale_min,
        )
        cv2.imwrite(str(review_dir / f"{path.stem}_review.png"), big)
        written += 1
    return written


def measure_scalebar_pixels(
    image_path: Path,
    *,
    top_fraction: float = 0.35,
    max_bar_height: int = 6,
    min_aspect: float = 4.0,
) -> int | None:
    """Measure the embedded scalebar bar width (px) near the top of an image.

    The exported brain crops carry a thin horizontal scalebar (e.g. a 35-px bar
    labelled "20 mm") near the top. Only the top ``top_fraction`` of rows is
    searched, so the brain (which sits lower and is tall) can never be mistaken
    for the bar. Among components there, the bar is the widest one that is thin
    (height <= ``max_bar_height``) and elongated (width/height >= ``min_aspect``).
    Uses the same inverted threshold the measurement pipeline binarises with.

    Returns the bar width in pixels, or ``None`` when no bar is found
    (e.g. full slices, which carry no embedded scalebar).
    """
    img = cv2.imread(str(image_path))
    if img is None:
        return None
    gray = cv2.cvtColor(img, cv2.COLOR_RGB2GRAY)  # match measurement_batch binarisation
    _, im_bw = cv2.threshold(gray, BINARY_THRESHOLD_DEFAULT, 255, 1)
    cutoff = max(1, int(round(im_bw.shape[0] * top_fraction)))
    num, _labels, stats, _ = cv2.connectedComponentsWithStats(im_bw[:cutoff, :], 8)
    best_width: int | None = None
    for i in range(1, num):  # label 0 is background
        w = int(stats[i, cv2.CC_STAT_WIDTH])
        h = int(stats[i, cv2.CC_STAT_HEIGHT])
        if h <= max_bar_height and w >= min_aspect * max(h, 1):
            if best_width is None or w > best_width:
                best_width = w
    return best_width


def detect_folder_pixel_size(
    image_files: list[Path],
    real_world_length: float,
    fallback_pixel_size: float,
) -> tuple[float, int | None]:
    """Pick a per-folder pixel size from the median embedded scalebar width.

    Each slice in a folder is rendered at the same zoom, so the bar width is
    constant within a folder; the median guards against the odd missed bar.
    Returns ``(pixel_size, median_bar_px)``, falling back to
    ``fallback_pixel_size`` (and ``None``) when no bar is detected anywhere.
    """
    widths = [w for w in (measure_scalebar_pixels(p) for p in image_files) if w]
    if not widths:
        return fallback_pixel_size, None
    median_px = int(round(float(np.median(widths))))
    return real_world_length / median_px, median_px


def single_pass_metrics(
    image_path: Path,
    pixel_size: float,
    kernel_size: int,
    cnt_threshold: float,
    area_close_mm: float = 0.0,
) -> tuple[float, float, float, float, float] | None:
    """Faithful single-pass recompute of the core morphometrics for one image.

    Mirrors the per-image math in ``measurement_batch.process_on_images_batch``
    (the ``compute_kernel_convex`` / ``compactness_2D`` helpers and the contour
    filter) but with two corrections: it drops the LGI threshold-climbing retry,
    which marches ``cnt_threshold`` past a small cropped brain and zeroes it out,
    and it segments with :func:`segment_foreground` instead of the grayscale
    brightness threshold, so bright labels (e.g. cyan) and the inner white cortex
    are no longer excluded from the region of interest.

    When ``area_close_mm > 0`` the reported ``area`` is measured on the
    morphologically-closed brain (option B: the white cortex ring between the
    rim and the core is filled in and counts as ROI), with the close distance
    scaled to ``pixel_size``. ``perimeter`` and ``lgi`` are always taken from the
    folded boundary, so the gyrification signal is unchanged either way.

    Returns ``(area, perimeter, perimeter_convex, lgi, compactness)`` in physical
    units, or ``None`` when no contour survives ``cnt_threshold``.
    """
    img = cv2.imread(str(image_path))
    if img is None:
        return None
    im_bw = segment_foreground(img)
    contours, _ = cv2.findContours(im_bw, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    kept = [c for c in contours if cv2.contourArea(c) > cnt_threshold]
    if not kept:
        return None

    kernel = compute_kernel_convex(kernel_size)
    inner_mask = np.zeros_like(im_bw)
    cv2.drawContours(inner_mask, kept, -1, 255, thickness=cv2.FILLED)
    closed = cv2.morphologyEx(inner_mask, cv2.MORPH_CLOSE, kernel)
    conv_contours, _ = cv2.findContours(closed, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    kept_conv = [c for c in conv_contours if cv2.contourArea(c) > cnt_threshold]

    # Perimeter and LGI from the folded boundary (gyrification).
    perimeter = sum(cv2.arcLength(c, True) for c in kept) * pixel_size
    perimeter_convex = (
        sum(cv2.arcLength(c, True) for c in kept_conv) * pixel_size if kept_conv else 0.0
    )
    lgi = perimeter / perimeter_convex if perimeter_convex else 0.0

    if area_close_mm > 0:
        # Option B: area on the cortex-ring-filled brain; compactness uses that
        # same region's boundary so it stays internally consistent.
        ak = area_close_kernel_px(area_close_mm, pixel_size, kernel_size)
        area_mask = closed_brain_mask(kept, im_bw.shape, ak)
        area_contours, _ = cv2.findContours(area_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        area_kept = [c for c in area_contours if cv2.contourArea(c) > cnt_threshold]
        area = sum(cv2.contourArea(c) for c in area_kept) * pixel_size ** 2
        area_perim = sum(cv2.arcLength(c, True) for c in area_kept) * pixel_size
        compactness = compactness_2D(area, area_perim)
    else:
        area = sum(cv2.contourArea(c) for c in kept) * pixel_size ** 2
        compactness = compactness_2D(area, perimeter)

    return area, perimeter, perimeter_convex, lgi, compactness


def single_pass_sulci(
    image_path: Path,
    pixel_size: float,
    cnt_threshold: float,
    unit: str = "mm",
) -> tuple[int, float | None, float | None, float | None, float | None]:
    """Recompute the sulci-depth statistics on the new segmentation.

    Mirrors the non-percent (cropped sub-slice) branch of
    ``measurement_batch.process_on_images_batch`` — convexity defects on the
    kept folded contours, kept when deeper than the fixed per-unit minimum — so
    the workbook's ``Sulci_count`` / depth columns agree with the markers drawn
    in :func:`render_review_image`. Returns ``(count, min, max, mean, total)``;
    depth stats are ``None`` when no defect is kept.
    """
    img = cv2.imread(str(image_path))
    if img is None:
        return 0, None, None, None, None
    mask = segment_foreground(img)
    contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    kept = [c for c in contours if cv2.contourArea(c) > cnt_threshold]
    min_keep = 0.5 if unit == "mm" else 0.05 if unit == "cm" else 0.0

    depths: list[float] = []
    for cnt in kept:
        hull = cv2.convexHull(cnt, returnPoints=False, clockwise=True)
        if hull is None or len(hull) < 3 or len(cnt) <= 3 or not np.all(np.diff(hull.ravel()) > 0):
            continue
        defects = cv2.convexityDefects(cnt, hull)
        if defects is None:
            continue
        for i in range(defects.shape[0]):
            depth_value = defects[i, 0][3] * pixel_size / DEFECT_FIXED_POINT
            if depth_value > min_keep:
                depths.append(depth_value)

    if not depths:
        return 0, None, None, None, None
    count = len(depths)
    return count, min(depths), max(depths), sum(depths) / count, sum(depths)


def correct_metrics_in_workbook(
    xlsx_path: Path,
    image_dir: Path,
    pixel_size: float,
    kernel_size: int,
    cnt_threshold: float,
    area_close_mm: float = 0.0,
) -> int:
    """Overwrite the core-metric cells of each data row with single-pass values.

    Looks up the area/perimeter/perimeter_convex/LGI/Compactness columns by
    header, then for every row whose ``File`` names an image in ``image_dir``
    recomputes those metrics via :func:`single_pass_metrics`. Rows whose
    ``File`` is not an image (the appended metadata rows) are left untouched.
    Returns the number of rows corrected.
    """
    metric_cols = ("area", "perimeter", "perimeter_convex", "LGI", "Compactness")
    wb = load_workbook(xlsx_path)
    ws = wb.active

    headers: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        if isinstance(value, str):
            headers[value.strip()] = col

    missing = [name for name in ("File", *metric_cols) if name not in headers]
    if missing:
        logging.warning("Skipping recompute for %s: missing columns %s", xlsx_path.name, missing)
        return 0

    file_col = headers["File"]
    corrected = 0
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=file_col).value
        if not isinstance(name, str):
            continue
        img_path = image_dir / name.strip()
        if not img_path.is_file():  # metadata rows ("PixelSize:", etc.) are not images
            continue
        result = single_pass_metrics(img_path, pixel_size, kernel_size, cnt_threshold, area_close_mm)
        values = result if result is not None else (0.0, 0.0, 0.0, 0.0, 0.0)
        if result is None:
            logging.warning("No contour > %s for %s; wrote zeros.", cnt_threshold, name)
        for metric_name, metric_value in zip(metric_cols, values):
            ws.cell(row=row, column=headers[metric_name], value=metric_value)
        corrected += 1

    wb.save(xlsx_path)
    return corrected


def correct_sulci_in_workbook(
    xlsx_path: Path,
    image_dir: Path,
    pixel_size: float,
    cnt_threshold: float,
    unit: str,
) -> int:
    """Overwrite Sulci_count and the depth columns with values recomputed on the
    new segmentation via :func:`single_pass_sulci`, so they agree with the area/
    LGI recompute and the review-image markers. Columns absent from the workbook
    are skipped. Returns the number of data rows updated.
    """
    count_col = "Sulci_count"
    depth_cols = {
        "min": f"min_depth_{unit}", "max": f"max_depth_{unit}",
        "mean": f"mean_depth_{unit}", "total": f"total_depth_{unit}",
    }
    wb = load_workbook(xlsx_path)
    ws = wb.active

    headers: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        if isinstance(value, str):
            headers[value.strip()] = col

    if "File" not in headers or count_col not in headers:
        logging.warning("Skipping sulci recompute for %s: missing File/%s", xlsx_path.name, count_col)
        return 0

    file_col = headers["File"]
    updated = 0
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=file_col).value
        if not isinstance(name, str):
            continue
        img_path = image_dir / name.strip()
        if not img_path.is_file():
            continue
        count, mn, mx, mean, total = single_pass_sulci(img_path, pixel_size, cnt_threshold, unit)
        ws.cell(row=row, column=headers[count_col], value=count)
        for key, header in depth_cols.items():
            if header not in headers:
                continue
            value = {"min": mn, "max": mx, "mean": mean, "total": total}[key]
            ws.cell(row=row, column=headers[header], value=value)
        updated += 1

    wb.save(xlsx_path)
    return updated


def is_metadata_row(first_cell: Any) -> bool:
    if not isinstance(first_cell, str):
        return False
    value = first_cell.strip()
    return value.endswith(":") or value in {"PixelSizeUnits", "KernelSize"}


def add_summary_table(xlsx_path: Path) -> None:
    wb = load_workbook(xlsx_path)
    ws = wb.active

    headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
    summary_start_col = ws.max_column + 3
    ws.cell(row=1, column=summary_start_col, value="Metric")
    ws.cell(row=1, column=summary_start_col + 1, value="Mean")

    data_rows: list[int] = []
    for row_idx in range(2, ws.max_row + 1):
        first_cell = ws.cell(row=row_idx, column=1).value
        if first_cell in (None, ""):
            continue
        if is_metadata_row(first_cell):
            continue
        data_rows.append(row_idx)

    skip_headers = {"File", "SliceKind"}
    summary_row = 2
    for col_idx, header in enumerate(headers, start=1):
        if header is None:
            continue
        header_text = str(header).strip()
        if header_text in skip_headers:
            continue

        values = []
        for row_idx in data_rows:
            value = ws.cell(row=row_idx, column=col_idx).value
            if value in (None, ""):
                continue
            try:
                values.append(float(value))
            except (TypeError, ValueError):
                continue

        if not values:
            continue

        ws.cell(row=summary_row, column=summary_start_col, value=header_text)
        mean_cell = ws.cell(
            row=summary_row,
            column=summary_start_col + 1,
            value=(sum(values) / len(values)),
        )
        mean_cell.number_format = "0.000000"
        summary_row += 1

    wb.save(xlsx_path)


def add_scalebar_metadata(
    xlsx_path: Path,
    *,
    measured_pixels: float | None,
    real_world_length: float | None,
    unit: str,
) -> None:
    """Append optional scale-from-scalebar metadata rows to the report workbook."""
    if measured_pixels is None or real_world_length is None:
        return
    wb = load_workbook(xlsx_path)
    ws = wb.active
    ws.append(["ScalebarMeasuredPixels:", measured_pixels])
    ws.append(["ScalebarRealWorldLength:", real_world_length])
    ws.append(["ScalebarRealWorldUnit:", unit])
    wb.save(xlsx_path)


def add_total_depth_column(xlsx_path: Path, unit: str) -> int:
    """Insert ``total_depth_<unit>`` as count * mean_depth for each data row."""
    wb = load_workbook(xlsx_path)
    ws = wb.active

    headers: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        if isinstance(value, str):
            headers[value.strip()] = col

    count_col = headers.get("Sulci_count")
    mean_col = headers.get(f"mean_depth_{unit}")
    total_name = f"total_depth_{unit}"
    if count_col is None or mean_col is None:
        logging.warning(
            "Skipping total-depth insertion for %s: missing Sulci_count or %s",
            xlsx_path.name,
            total_name,
        )
        return 0

    total_col = headers.get(total_name)
    if total_col is None:
        insert_at = mean_col
        ws.insert_cols(insert_at)
        ws.cell(row=1, column=insert_at, value=total_name)
        total_col = insert_at
        if mean_col >= insert_at:
            mean_col += 1

    updated = 0
    for row_idx in range(2, ws.max_row + 1):
        first_cell = ws.cell(row=row_idx, column=1).value
        if first_cell in (None, "") or is_metadata_row(first_cell):
            continue
        count_value = ws.cell(row=row_idx, column=count_col).value
        mean_value = ws.cell(row=row_idx, column=mean_col).value
        if count_value in (None, "") or mean_value in (None, ""):
            ws.cell(row=row_idx, column=total_col, value=None)
            continue
        try:
            total_value = float(count_value) * float(mean_value)
        except (TypeError, ValueError):
            ws.cell(row=row_idx, column=total_col, value=None)
            continue
        cell = ws.cell(row=row_idx, column=total_col, value=total_value)
        cell.number_format = "0.000000"
        updated += 1

    wb.save(xlsx_path)
    return updated


def run_folder(
    input_dir: Path,
    output_dir: Path,
    week: int,
    axis: str,
    *,
    pixel_size: float,
    scalebar_measured_pixels: float | None,
    scalebar_real_world_length: float | None,
    kernel_size: int,
    cnt_threshold: float,
    unit: str,
    auto_scalebar: bool = False,
    single_pass_metrics_enabled: bool = False,
    review: bool = False,
    review_upscale_min: int = 600,
    area_close_mm: float = 0.0,
) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)

    # Per-folder calibration: prefer the images' own embedded scalebar over the
    # single configured pixel size, since each week/axis is rendered at its own
    # zoom. Falls back to the configured pixel_size when no bar is detected.
    eff_pixel_size = pixel_size
    detected_bar_px: int | None = None
    if auto_scalebar and scalebar_real_world_length:
        eff_pixel_size, detected_bar_px = detect_folder_pixel_size(
            find_image_files(input_dir), scalebar_real_world_length, pixel_size
        )
        if detected_bar_px:
            logging.info(
                "Auto-scalebar week=%s axis=%s: %d px = %.4f %s -> %.6f %s/pixel",
                week, axis, detected_bar_px, scalebar_real_world_length, unit,
                eff_pixel_size, unit,
            )
        else:
            logging.warning(
                "No scalebar detected in %s; using configured %.6f %s/pixel",
                input_dir, eff_pixel_size, unit,
            )

    process_on_images_batch(
        str(input_dir),
        str(output_dir),
        pixel_size=eff_pixel_size,
        kernel_size=kernel_size,
        cnt_threshold=cnt_threshold,
        unit=unit,
    )

    default_xlsx = output_dir / "Batch_Allmarks.xlsx"
    final_xlsx = output_dir / f"week{week}_{axis}_Batch_Allmarks.xlsx"
    if final_xlsx.exists():
        final_xlsx.unlink()
    default_xlsx.replace(final_xlsx)
    add_scalebar_metadata(
        final_xlsx,
        measured_pixels=(float(detected_bar_px) if detected_bar_px else scalebar_measured_pixels),
        real_world_length=scalebar_real_world_length,
        unit=unit,
    )
    n_total = add_total_depth_column(final_xlsx, unit)
    logging.info("Added/updated total-depth values in %d data row(s) in %s", n_total, final_xlsx.name)

    # Recompute the core metrics single-pass BEFORE the summary table is built,
    # so the appended means reflect the corrected (retry-free) values.
    if single_pass_metrics_enabled:
        n = correct_metrics_in_workbook(
            final_xlsx, input_dir, eff_pixel_size, kernel_size, cnt_threshold, area_close_mm
        )
        logging.info("Single-pass recompute corrected %d data row(s) in %s", n, final_xlsx.name)
        n_sulci = correct_sulci_in_workbook(
            final_xlsx, input_dir, eff_pixel_size, cnt_threshold, unit
        )
        logging.info("Sulci recompute (new segmentation) updated %d data row(s) in %s",
                     n_sulci, final_xlsx.name)

    add_summary_table(final_xlsx)

    if review:
        n_review = render_review_images(
            input_dir, output_dir, kernel_size, cnt_threshold, eff_pixel_size, unit,
            area_close_mm, upscale_min=review_upscale_min,
        )
        logging.info("Wrote %d enlarged review image(s) to %s", n_review, output_dir / "review")

        # The corrected review/ images supersede the core function's old-method
        # annotated PNGs in image_Batch/, so drop that folder to avoid confusion.
        # (Only when review is on; otherwise image_Batch is the sole image output.)
        batch_dir = output_dir / "image_Batch"
        if batch_dir.is_dir():
            shutil.rmtree(batch_dir, ignore_errors=True)
            logging.info("Removed stale old-method annotations: %s", batch_dir)

    return final_xlsx


def main() -> int:
    args = parse_args()
    logging.basicConfig(level=getattr(logging, args.log_level), format="%(levelname)s: %(message)s")
    settings = resolve_settings(args)

    input_root: Path = settings["input_root"]
    output_root: Path = settings["output_root"]
    section_label: str = settings["section_label"]
    weeks: list[int] = settings["weeks"]
    axes: list[str] = settings["axes"]

    if not input_root.exists():
        logging.error("Input root does not exist: %s", input_root)
        return 1

    output_root.mkdir(parents=True, exist_ok=True)

    if settings["auto_scalebar"]:
        logging.info(
            "Auto-scalebar ON: pixel size measured per folder from the embedded "
            "%.6f %s scalebar (configured %.6f %s/pixel used only as fallback).",
            settings["scalebar_real_world_length"],
            settings["unit"],
            settings["pixel_size"],
            settings["unit"],
        )
    elif settings["scalebar_measured_pixels"] is not None:
        logging.info(
            "Using scalebar calibration: %.6f %s/pixel from %.6f px = %.6f %s",
            settings["pixel_size"],
            settings["unit"],
            settings["scalebar_measured_pixels"],
            settings["scalebar_real_world_length"],
            settings["unit"],
        )
    else:
        logging.info(
            "Using direct pixel size: %.6f %s/pixel",
            settings["pixel_size"],
            settings["unit"],
        )
    if settings["single_pass_metrics"]:
        logging.info(
            "Single-pass metrics ON: area/perimeter/perimeter_convex/LGI/Compactness "
            "recomputed without the LGI retry after each batch run."
        )
    if settings["area_close_mm"] > 0:
        logging.info(
            "Option B ON: area measured on the brain after a %.1f %s morphological "
            "close (cortex ring filled, scaled per folder); LGI/perimeter stay on "
            "the folded boundary.",
            settings["area_close_mm"], settings["unit"],
        )

    processed = 0
    skipped = 0
    failed = 0

    for week in weeks:
        for axis in axes:
            input_dir = input_root / str(week) / axis
            output_dir = output_root / str(week) / section_label / axis

            if not input_dir.exists():
                logging.warning("Missing folder, skipping: %s", input_dir)
                skipped += 1
                continue

            if not find_image_files(input_dir):
                logging.warning("No images found, skipping: %s", input_dir)
                skipped += 1
                continue

            logging.info("Processing week=%s axis=%s from %s", week, axis, input_dir)
            try:
                xlsx_path = run_folder(
                    input_dir,
                    output_dir,
                    week,
                    axis,
                    pixel_size=settings["pixel_size"],
                    scalebar_measured_pixels=settings["scalebar_measured_pixels"],
                    scalebar_real_world_length=settings["scalebar_real_world_length"],
                    kernel_size=settings["kernel_size"],
                    cnt_threshold=settings["cnt_threshold"],
                    unit=settings["unit"],
                    auto_scalebar=settings["auto_scalebar"],
                    single_pass_metrics_enabled=settings["single_pass_metrics"],
                    review=settings["review"],
                    review_upscale_min=settings["review_upscale_min"],
                    area_close_mm=settings["area_close_mm"],
                )
                logging.info("Wrote workbook: %s", xlsx_path)
                processed += 1
            except Exception as exc:
                logging.exception("Failed week=%s axis=%s: %s", week, axis, exc)
                failed += 1

    logging.info("Done. processed=%s skipped=%s failed=%s", processed, skipped, failed)
    return 1 if failed else 0


if __name__ == "__main__":
    raise SystemExit(main())
