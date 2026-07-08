"""Shared writer / reader for FetoMorph result Excel files.

Lays out one or more sheets in the same visual structure as the
``Concise`` sheet of ``data_collection_plan.xlsx``:

    Results
        File name / Folder / User / Date

    Mean results
        Section | Area | Perimeter | LGI | Compactness |
        Primary/Secondary/Tertiary/UnclassifiedSulciCount |
        Primary/Secondary/Tertiary/UnclassifiedMeanDepth

    Totals (optional)
        Volume / Surface Area / GI / Compactness / Total sulci / Mean depth

    Parameters
        Kernel size (mm) / Kernel size (px) / Pixel spacing / Slice thickness /
        Filtered threshold / ...

    Footer: "The results were produced by FetoMorph."

All measurement entry points (NIfTI, STL, VTK allmarks, and the dock
``Export Excel…``) build a :class:`ResultsSheet` and hand it to
:func:`write_results_workbook`.
"""

from __future__ import annotations

import getpass
import math
import os
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


RESULTS_COLUMNS: tuple[str, ...] = (
    "Section",
    "Area",
    "Perimeter",
    "Perimeter_interior",
    "LGI",
    "Compactness",
    "PrimarySulciCount",
    "SecondarySulciCount",
    "TertiarySulciCount",
    "UnclassifiedSulciCount",
    "PrimaryMeanDepth",
    "SecondaryMeanDepth",
    "TertiaryMeanDepth",
    "UnclassifiedMeanDepth",
)

TITLE = "Results"
PARAMETERS_HEADER = "Parameters"
RESULTS_HEADER = "Mean results"
TOTALS_HEADER = "Totals"
FOOTER = "The results were produced by FetoMorph."

PARAMETER_KEYS = (
    "Kernel size (mm)",
    "Kernel size (px)",
    "Pixel spacing",
    "Slice thickness",
    "Filtered threshold",
)


@dataclass
class ResultsSheet:
    """One sheet's worth of metric output."""

    sheet_name: str
    file_name: str | None = None
    folder: str | None = None
    user: str | None = None
    date: str | None = None
    parameters: dict[str, Any] = field(default_factory=dict)
    rows: list[dict[str, Any]] = field(default_factory=list)
    totals: dict[str, Any] | None = None
    # Optional explanatory note written one cell to the right of a totals
    # value, keyed by the same totals key.
    totals_notes: dict[str, Any] = field(default_factory=dict)
    embed_section_images: bool = True
    image_max_width: int = 900
    # Extra column headers inserted into the Mean results table between
    # ``Section`` and ``Area``. Use this when each row needs to carry its
    # own per-run parameters (kernel size, pixel spacing, threshold, …)
    # because they vary across measurements of the same file.
    extra_columns: tuple[str, ...] = ()
    # Extra column headers appended AFTER the metric columns (to the right of the
    # RESULTS_COLUMNS block). Use for variable-length per-row detail that belongs
    # at the end of the row, e.g. every individual sulcus depth per class. Like
    # every other column they are dropped when ``drop_empty_columns`` and empty.
    trailing_columns: tuple[str, ...] = ()
    # When True, columns whose every row value is empty get dropped from
    # the Mean results table, and parameter rows with no value get
    # dropped from the Parameters block (skipping the block entirely if
    # nothing is left). ``Section`` is always preserved.
    drop_empty_columns: bool = False
    # Header for the first (row-label) column. Defaults to "Section"; the
    # cross-measurement dock export sets it to "File name".
    section_header: str = "Section"
    # File kind (e.g. "Area", "Volume"). When set, the header row shows a
    # single "kind" field in place of the File name / Folder pair — used by
    # the cross-measurement dock export, where each sheet is one kind and the
    # per-file name / folder live in the table below.
    kind: str | None = None


def subtype_mean(direct_mean: Any, v_values: Iterable[Any]) -> float | None:
    """Pick the per-subtype mean depth, averaging v1/v2/v3 when needed.

    Mirrors the rule the per-slice exporters use:
    when a sulcus subtype has 3 or fewer sulci on a slice, the raw
    values land in ``Primary_v1_unit`` / ``_v2_`` / ``_v3_`` and the
    ``_mean_`` column stays empty. When count > 3, the ``_mean_`` cell
    is populated and the v-cells are empty.
    """
    if _is_real_number(direct_mean):
        return float(direct_mean)
    floats = [float(v) for v in v_values if _is_real_number(v)]
    if not floats:
        return None
    return sum(floats) / len(floats)


def build_measurement_sheet(
    file_path: str,
    hallmark: str,
    rows: list[dict[str, Any]],
    parameters: dict[str, Any] | None = None,
    totals: dict[str, Any] | None = None,
    *,
    extra_columns: tuple[str, ...] = (),
    totals_notes: dict[str, Any] | None = None,
) -> "ResultsSheet":
    """Build a spec-layout ``ResultsSheet`` for a single measurement tool.

    Gives the standalone Area / Volume / LGI / Sulci / Compactness exports the
    same Excel style as All-hallmarks. ``hallmark`` (e.g. "Area", "Volume",
    "All") is written as the first entry of the Totals block ("Hallmark
    computed"). Empty columns are dropped automatically.
    """
    totals_out: dict[str, Any] = {"Hallmark computed": hallmark}
    totals_out.update(totals or {})
    return ResultsSheet(
        sheet_name=os.path.basename(file_path) or "Results",
        file_name=os.path.basename(file_path) or None,
        folder=os.path.dirname(file_path) or None,
        parameters=parameters or {},
        rows=rows or [],
        extra_columns=tuple(extra_columns),
        totals=totals_out,
        totals_notes=totals_notes or {},
        drop_empty_columns=True,
    )


def gi_3d_note(gi_3d: Any) -> str:
    """One-line explanation for the 3-D area-based GI (convex-hull) cell."""
    if not _is_real_number(gi_3d):
        return ""
    g = float(gi_3d)
    return (
        f"3-D gyrification = exact mesh surface ÷ convex-hull surface "
        f"= {g:.2f}. ≈1 for a smooth/convex surface; larger means more folding "
        f"(FreeSurfer-style 3-D GI proxy). Independent of the 2-D perimeter GI."
    )


def write_results_workbook(
    xlsx_path: str,
    sheets: list[ResultsSheet],
) -> str:
    """Write *sheets* to *xlsx_path* using the spec layout.

    Returns the saved path.
    """
    if not sheets:
        raise ValueError("write_results_workbook: no sheets given")

    wb = Workbook()
    wb.remove(wb.active)

    xlsx_dir = os.path.dirname(os.path.abspath(xlsx_path))
    used_names: set[str] = set()
    for sheet in sheets:
        name = _safe_sheet_name(sheet.sheet_name or "Results", used_names)
        used_names.add(name)
        ws = wb.create_sheet(title=name)
        _render_sheet(wb, ws, sheet, used_names, xlsx_dir=xlsx_dir)

    os.makedirs(xlsx_dir or ".", exist_ok=True)
    wb.save(xlsx_path)
    return xlsx_path


def read_results_sheet(
    xlsx_path: str,
    *,
    sheet_index: int = 0,
    sheet_name: str | None = None,
) -> dict[str, Any]:
    """Read back a sheet written by :func:`write_results_workbook`.

    Returns a dict with keys ``file_name``, ``folder``, ``user``, ``date``,
    ``parameters`` (dict), ``rows`` (list of dicts keyed by header), and
    ``totals`` (dict, possibly empty).
    """
    wb = load_workbook(xlsx_path, data_only=True, read_only=False)
    ws = (wb[sheet_name] if sheet_name is not None
          else wb.worksheets[sheet_index])
    cells = [[c.value for c in row] for row in ws.iter_rows()]

    out: dict[str, Any] = {
        "file_name": None,
        "folder": None,
        "kind": None,
        "user": None,
        "date": None,
        "parameters": {},
        "rows": [],
        "totals": {},
    }

    section: str | None = None
    table_headers: list[str] | None = None
    for raw in cells:
        if not raw or all(v is None for v in raw):
            section = None
            table_headers = None
            continue

        label = raw[1] if len(raw) >= 2 else None
        if isinstance(label, str):
            if label.strip() == TITLE:
                continue
            if label.strip() == PARAMETERS_HEADER:
                section = "parameters"
                continue
            if label.strip() == RESULTS_HEADER:
                section = "results"
                table_headers = None
                continue
            if label.strip() == TOTALS_HEADER:
                section = "totals"
                continue
            if label.strip() == FOOTER:
                section = None
                continue

        # Row-3 metadata pairs (B/C, D/E, F/G, H/I)
        if _stringly_equals(raw[1] if len(raw) > 1 else None, "File name"):
            out["file_name"] = raw[2] if len(raw) > 2 else None
            out["folder"] = raw[4] if len(raw) > 4 else None
            out["user"] = raw[6] if len(raw) > 6 else None
            out["date"] = raw[8] if len(raw) > 8 else None
            continue

        # Dock-export variant: "kind" replaces the File name / Folder pair.
        if _stringly_equals(raw[1] if len(raw) > 1 else None, "kind"):
            out["kind"] = raw[2] if len(raw) > 2 else None
            out["user"] = raw[4] if len(raw) > 4 else None
            out["date"] = raw[6] if len(raw) > 6 else None
            continue

        if section == "parameters":
            key = label
            value = raw[2] if len(raw) >= 3 else None
            if isinstance(key, str) and key.strip():
                out["parameters"][key.strip()] = value
        elif section == "results":
            if table_headers is None:
                table_headers = [(str(v).strip() if v is not None else "")
                                 for v in raw[1:]]
            else:
                row = {}
                for i, header in enumerate(table_headers):
                    if not header:
                        continue
                    cell_index = i + 1
                    row[header] = (raw[cell_index]
                                   if cell_index < len(raw) else None)
                out["rows"].append(row)
        elif section == "totals":
            key = label
            value = raw[2] if len(raw) >= 3 else None
            if isinstance(key, str) and key.strip():
                out["totals"][key.strip()] = value

    return out


# ---------------------------------------------------------------------
# Internals
# ---------------------------------------------------------------------


_TITLE_FONT = Font(bold=True, size=14, color="FFFFFF")
_TITLE_FILL = PatternFill("solid", fgColor="1F6FA5")
_SECTION_FONT = Font(bold=True, size=11, color="FFFFFF")
_SECTION_FILL = PatternFill("solid", fgColor="2C3E50")
_LABEL_FONT = Font(bold=True)
_TABLE_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TABLE_HEADER_FILL = PatternFill("solid", fgColor="3498DB")
_FOOTER_FONT = Font(italic=True, color="555555")
_LINK_FONT = Font(color="0563C1", underline="single")
_CENTER = Alignment(horizontal="center")

SECTION_LINK_KEY = "_section_link"


# One-line explanation rendered in the cell next to each totals value. Matched
# by substring against the totals key (which carries units), most-specific first
# so e.g. "GI 3D" wins over "GI" and "LGI". Keep these in sync with the totals
# labels produced by the measurement exporters (VTK is the canonical style).
_TOTAL_NOTES = (
    # --- gyrification (3-D) ---
    ("GI 3D", "3-D gyrification index = exact mesh surface ÷ convex-hull (smooth envelope) surface."),
    ("Convex hull area", "Surface area of the mesh's convex hull (the smooth outer envelope)."),
    # --- surface-area breakdown (lowercase 'area' is never caught by 'Surface Area') ---
    ("Area lateral surface", "Lateral surface = ∫ exterior perimeter dh (Simpson), excluding end caps."),
    ("Surface area lateral", "Lateral surface = ∫ exterior perimeter dh (Simpson), excluding end caps."),
    ("Surface area caps", "Top + bottom end-face (cap) areas, added to the lateral surface."),
    ("Surface Area", "Total 3-D surface area = lateral (∫ perimeter dh) + top & bottom caps."),
    # --- volume ---
    ("Volume", "Tissue volume = ∫ cross-section area dh (Simpson integration over slices)."),
    # --- compactness (batch 2-D mean before the generic 3-D note) ---
    ("Compactness (mean across images)", "Mean 2-D compactness (4π·A ÷ P²) across all processed images."),
    ("Compactness", "3-D compactness / sphericity = 36π·V² ÷ S³ (1 = a perfect sphere)."),
    # --- cavity correction ---
    ("Surface connected cavity area", "Area of surface-connected (open) cavities removed from the volume."),
    ("Cavity area removed", "Area of surface-connected (open) cavities removed from the volume."),
    ("Cavity wall surface", "Wall area of open cavities, added to the 3-D surface area."),
    ("Cavity wall perimeter added", "Wall perimeter of open cavities, added to the lateral surface."),
    ("Number of surface-connected cavities", "Open cavities corrected: area removed from volume, walls added to surface."),
    ("Number of enclosed cavities", "Fully-enclosed voids left as solid (not corrected)."),
    # --- sulci ---
    ("Total sulci count", "Number of detected sulci (convexity defects passing the depth filter)."),
    ("Min sulci depth", "Smallest depth among all detected sulci."),
    ("Max sulci depth", "Largest depth among all detected sulci."),
    ("Mean sulci depth", "Mean depth across all detected sulci."),
    # --- gyrification (2-D perimeter ratio; keep last so 'GI 3D' / 'LGI' win) ---
    ("LGI (mean across images)", "Mean local gyrification index (exterior ÷ closed-envelope perimeter) across all images."),
    ("GI", "Gyrification index = Σ exterior (cortical) perimeter ÷ Σ closed-envelope perimeter."),
)


def _default_total_note(key) -> str:
    """Return a one-line explanation for a totals key (matched by substring)."""
    k = str(key)
    for sub, note in _TOTAL_NOTES:
        if sub in k:
            return note
    return ""


def _render_sheet(wb, ws, sheet: ResultsSheet, used_names: set[str],
                  *, xlsx_dir: str = "") -> None:
    user = sheet.user or _safe_username()
    date_str = sheet.date or datetime.now().strftime("%Y-%m-%d %H:%M")
    file_name = sheet.file_name or ""
    folder = sheet.folder or ""
    n_cols = max(
        len(RESULTS_COLUMNS) + len(sheet.extra_columns or ())
        + len(sheet.trailing_columns or ()) + 1, 11)
    last_col = n_cols + 1  # column A stays empty; data starts at B

    # Row 1 — Title
    cell = ws.cell(row=1, column=2, value=TITLE)
    cell.font = _TITLE_FONT
    cell.fill = _TITLE_FILL
    ws.merge_cells(start_row=1, start_column=2,
                   end_row=1, end_column=last_col)

    # Row 3 — metadata pairs. When a kind is set (cross-measurement dock
    # export) the header reads "kind" and the per-file name / folder are
    # carried by the table below; otherwise it keeps File name / Folder.
    if sheet.kind:
        pairs = (
            ("kind", sheet.kind),
            ("User", user),
            ("Date", date_str),
        )
    else:
        pairs = (
            ("File name", file_name),
            ("Folder", folder),
            ("User", user),
            ("Date", date_str),
        )
    col = 2
    for label, value in pairs:
        c = ws.cell(row=3, column=col, value=label)
        c.font = _LABEL_FONT
        ws.cell(row=3, column=col + 1, value=value)
        col += 2

    # Parameters and Totals are placed at the TOP (before the per-section table)
    # so the run settings and the summary numbers read first.
    row = 5

    # Parameters block
    params = _sorted_parameters(sheet.parameters)
    if sheet.drop_empty_columns:
        params = [(k, v) for (k, v) in params if _is_populated(v)]
    elif not params:
        params = [(k, None) for k in PARAMETER_KEYS]
    if params:
        _write_section(ws, row, last_col, PARAMETERS_HEADER)
        row += 1
        for key, value in params:
            ws.cell(row=row, column=2, value=key).font = _LABEL_FONT
            ws.cell(row=row, column=3, value=_to_cell(value))
            row += 1
        row += 1  # blank line

    # Totals block — each total carries a one-line explanation in the next cell.
    totals_items = list(sheet.totals.items()) if sheet.totals else []
    if sheet.drop_empty_columns:
        totals_items = [(k, v) for k, v in totals_items if _is_populated(v)]
    if totals_items:
        _write_section(ws, row, last_col, TOTALS_HEADER)
        row += 1
        notes = sheet.totals_notes or {}
        for key, value in totals_items:
            ws.cell(row=row, column=2, value=key).font = _LABEL_FONT
            ws.cell(row=row, column=3, value=_to_cell(value))
            note = notes.get(key) or _default_total_note(key)
            if note:
                nc = ws.cell(row=row, column=4, value=str(note))
                nc.font = _FOOTER_FONT
                nc.alignment = Alignment(horizontal="left")
            row += 1
        row += 1

    # Mean results header
    _write_section(ws, row, last_col, RESULTS_HEADER)
    row += 1

    # Column header row. Per-row parameter columns (when the caller
    # set ``extra_columns``) sit between Section and the metric columns
    # so each measurement's adjustment parameters travel next to its row.
    extras = tuple(sheet.extra_columns or ())
    trailing = tuple(sheet.trailing_columns or ())
    column_names = ((sheet.section_header,) + extras
                    + RESULTS_COLUMNS[1:] + trailing)
    if sheet.drop_empty_columns:
        column_names = _filter_populated_columns(column_names, sheet.rows)
    for i, name in enumerate(column_names, start=2):
        c = ws.cell(row=row, column=i, value=name)
        c.font = _TABLE_HEADER_FONT
        c.fill = _TABLE_HEADER_FILL
        c.alignment = _CENTER
    row += 1

    # Data rows. If the row carries a SECTION_LINK_KEY pointing to a
    # local image, embed the image on a per-section tab and link the
    # Section cell internally — Excel does not raise the "external
    # link / security" prompt for in-document navigation. URLs and
    # non-image links fall back to external hyperlinks.
    for r in sheet.rows:
        link_target = r.get(SECTION_LINK_KEY)
        for i, name in enumerate(column_names, start=2):
            cell = ws.cell(row=row, column=i, value=_to_cell(r.get(name)))
            if name == sheet.section_header and link_target:
                _wire_section_link(
                    wb, cell, link_target,
                    main_sheet=ws,
                    main_sheet_name=ws.title,
                    section_value=r.get(sheet.section_header),
                    used_names=used_names,
                    embed_images=sheet.embed_section_images,
                    image_max_width=sheet.image_max_width,
                    xlsx_dir=xlsx_dir,
                )
        row += 1
    row += 1  # blank line

    # Footer
    cell = ws.cell(row=row, column=2, value=FOOTER)
    cell.font = _FOOTER_FONT
    cell.alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=row, start_column=2,
                   end_row=row, end_column=last_col)

    # Column widths
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    for i in range(4, last_col + 1):
        ws.column_dimensions[get_column_letter(i)].width = 16


def _write_section(ws, row: int, last_col: int, label: str) -> None:
    c = ws.cell(row=row, column=2, value=label)
    c.font = _SECTION_FONT
    c.fill = _SECTION_FILL
    ws.merge_cells(start_row=row, start_column=2,
                   end_row=row, end_column=last_col)


def _sorted_parameters(params: dict[str, Any]) -> list[tuple[str, Any]]:
    """Keep the canonical PARAMETER_KEYS order, then any extras."""
    if not params:
        return []
    ordered: list[tuple[str, Any]] = []
    seen: set[str] = set()
    for key in PARAMETER_KEYS:
        if key in params:
            ordered.append((key, params[key]))
            seen.add(key)
    for k, v in params.items():
        if k not in seen:
            ordered.append((k, v))
    return ordered


def _safe_sheet_name(name: str, used: set[str]) -> str:
    """Excel sheet names: ≤31 chars, no []*?/\\:."""
    cleaned = "".join("_" if ch in r"[]*?/\:" else ch for ch in str(name))
    cleaned = cleaned.strip() or "Results"
    cleaned = cleaned[:31]
    if cleaned not in used:
        return cleaned
    base = cleaned[:28]
    for i in range(2, 1000):
        candidate = f"{base}_{i}"
        if candidate not in used:
            return candidate
    return cleaned


def _to_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, str)):
        return value
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return None
        return value
    try:
        if isinstance(value, datetime):
            return value
    except Exception:
        pass
    return str(value)


def _is_populated(value: Any) -> bool:
    """Return False for None / NaN / empty strings — these are the
    cells that ``drop_empty_columns`` is allowed to remove."""
    if value is None:
        return False
    if isinstance(value, float):
        try:
            return not math.isnan(value)
        except Exception:
            return False
    if isinstance(value, str):
        return value.strip() != ""
    return True


def _filter_populated_columns(column_names: tuple[str, ...],
                              rows: list[dict[str, Any]]) -> tuple[str, ...]:
    """Drop columns whose every row value is empty.

    ``Section`` is always preserved so each row still has a label even
    when every other column is empty.
    """
    if not rows:
        return column_names
    kept: list[str] = []
    for idx, name in enumerate(column_names):
        if idx == 0:  # row-label column (Section / File name) is always kept
            kept.append(name)
            continue
        if any(_is_populated(r.get(name)) for r in rows):
            kept.append(name)
    return tuple(kept)


def _safe_username() -> str:
    try:
        return getpass.getuser()
    except Exception:
        return os.environ.get("USER") or os.environ.get("USERNAME") or ""


def _is_real_number(v: Any) -> bool:
    if v is None or isinstance(v, bool):
        return False
    if isinstance(v, (int, float)):
        try:
            return not math.isnan(float(v))
        except Exception:
            return False
    if isinstance(v, str):
        try:
            float(v)
            return True
        except ValueError:
            return False
    return False


def _stringly_equals(value: Any, target: str) -> bool:
    return isinstance(value, str) and value.strip() == target


def _link_href(link: Any, xlsx_dir: str) -> str | None:
    """Build the Excel hyperlink string for *link*.

    Local paths are rewritten relative to *xlsx_dir* so the workbook +
    slices folder stay portable when copied with ``Save Data``. Existing
    URL-style strings (``http://…``, ``file://…``) are passed through.
    """
    if not isinstance(link, str) or not link:
        return None
    if "://" in link:
        return link
    if not xlsx_dir:
        return link
    try:
        abspath = os.path.abspath(link)
        rel = os.path.relpath(abspath, start=xlsx_dir)
        return rel.replace(os.sep, "/")
    except Exception:
        return link


_IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".tif", ".tiff"}


def _wire_section_link(wb, cell, link_target, *, main_sheet,
                       main_sheet_name, section_value, used_names,
                       embed_images: bool, image_max_width: int,
                       xlsx_dir: str) -> None:
    """Attach a hyperlink to *cell* pointing at the section's slice.

    When the target is a readable local image and *embed_images* is
    True, the image is embedded on its own per-section sheet and the
    Section cell becomes an internal navigation link (Excel skips the
    external-link security prompt for in-document destinations). A
    "Back to results" link on the image sheet returns to *main_sheet*.
    Falls back to an external relative hyperlink if anything goes
    wrong (file missing, PIL unavailable, etc.).
    """
    cell.font = _LINK_FONT
    is_image = (isinstance(link_target, str)
                and os.path.splitext(link_target)[1].lower() in _IMAGE_EXTS
                and os.path.isfile(link_target))
    if embed_images and is_image:
        try:
            image_sheet_name = _embed_image_sheet(
                wb, link_target,
                section_value=section_value,
                main_sheet_name=main_sheet_name,
                used_names=used_names,
                max_width=image_max_width,
            )
        except Exception as ex:
            print(f"[ResultsExcel] WARN: image embed failed for "
                  f"{link_target}: {ex}; falling back to external link.")
            image_sheet_name = None
        if image_sheet_name:
            cell.hyperlink = f"#'{image_sheet_name}'!A1"
            return

    href = _link_href(link_target, xlsx_dir)
    if href:
        cell.hyperlink = href


def _embed_image_sheet(wb, image_path: str, *, section_value,
                       main_sheet_name: str,
                       used_names: set[str],
                       max_width: int) -> str:
    """Create a per-section sheet with the embedded image and a Back link.

    Returns the resulting sheet name.
    """
    from openpyxl.drawing.image import Image as XlImage

    base = f"Slice_{section_value}" if section_value is not None else "Slice"
    sheet_name = _safe_sheet_name(base, used_names)
    used_names.add(sheet_name)

    img_ws = wb.create_sheet(title=sheet_name)
    back = img_ws.cell(row=1, column=1, value="← Back to results")
    back.font = _LINK_FONT
    back.hyperlink = f"#'{main_sheet_name}'!A1"

    img_ws.cell(row=2, column=1,
                value=os.path.basename(image_path)).font = Font(italic=True,
                                                                color="555555")

    # Optionally downscale the image so the embedded copy doesn't bloat
    # the workbook. The original PNG on disk is untouched.
    embed_path = _maybe_thumbnail(image_path, max_width=max_width)
    img = XlImage(embed_path)
    img_ws.add_image(img, "A4")
    img_ws.sheet_view.showGridLines = False
    img_ws.column_dimensions["A"].width = 30
    return sheet_name


def _maybe_thumbnail(image_path: str, *, max_width: int) -> str:
    """If the image is wider than *max_width*, return a thumbnail copy.

    The thumbnail is written to a process-local temp directory so the
    user's slice folder stays clean. openpyxl reads the file at
    ``wb.save()`` time and embeds the bytes into the workbook, so the
    on-disk thumbnail can be discarded afterwards (handled by the OS
    when ``tempfile.gettempdir()`` is cleaned). Falls back to the
    original path if PIL is missing or anything else goes wrong.
    """
    try:
        from PIL import Image as PilImage
    except Exception:
        return image_path
    try:
        with PilImage.open(image_path) as im:
            if im.width <= max_width:
                return image_path
            ratio = max_width / float(im.width)
            new_size = (max_width, max(1, int(im.height * ratio)))
            thumb = im.convert("RGB").resize(new_size, PilImage.LANCZOS)
        import tempfile
        base = os.path.splitext(os.path.basename(image_path))[0]
        thumb_dir = os.path.join(tempfile.gettempdir(),
                                 "fetomorph_thumbs")
        os.makedirs(thumb_dir, exist_ok=True)
        thumb_path = os.path.join(
            thumb_dir, f"{base}_{hash(image_path) & 0xFFFFFFFF:x}.png")
        thumb.save(thumb_path, format="PNG", optimize=True)
        return thumb_path
    except Exception:
        return image_path
