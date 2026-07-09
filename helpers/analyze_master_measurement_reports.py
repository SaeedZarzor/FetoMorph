"""Analyze the workbooks produced by ``process_on_images_batch``.

Each ``*Batch_Allmarks.xlsx`` written by :func:`functions.measurement_batch.
process_on_images_batch` uses the shared spec layout (Results / Parameters /
Totals / Mean results / footer) rather than a flat table, so this reader parses
it with :func:`helpers.results_excel_format.read_results_sheet` instead of
``pandas.read_excel``.

The spec-layout "Mean results" table carries, per slice:

* core metrics — ``Area`` / ``Perimeter`` / ``LGI`` / ``Compactness``;
* per-class counts — ``PrimarySulciCount`` … ``UnclassifiedSulciCount``;
* per-class mean depth — ``PrimaryMeanDepth`` … ``UnclassifiedMeanDepth``;
* every individual sulcus depth — ``Primary_depth_1`` … ``Unclassified_depth_N``
  (deepest first; blanks past a slice's own count).

Because the real per-sulcus depths are now columns in their own right, the old
count+min/max/mean reconstruction is gone: this script reads the values
directly. The per-class and overall depth summaries also report full
``normalized_{mean,std,min,max}`` stats over the per-sulcus normalized depths
(each depth / its image's deepest sulcus), read from the ``{Class}_depth_i_norm``
columns written by ``process_on_images_batch``.
For each workbook it appends an ``Analysis`` sheet (summary tables +
boxplots) back into the same file. Cropped sub-slices only carry the
``Unclassified`` class; full slices carry primary/secondary/tertiary too.
"""

from __future__ import annotations

import argparse
import io
import math
import re
import sys
from pathlib import Path
from typing import Iterable

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parent
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from helpers.results_excel_format import read_results_sheet

try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
except Exception as exc:  # pragma: no cover - environment-dependent import failure
    plt = None
    MATPLOTLIB_IMPORT_ERROR = exc
else:
    MATPLOTLIB_IMPORT_ERROR = None

if plt is not None:
    from openpyxl.drawing.image import Image as XLImage
else:
    XLImage = None


# Column names in the spec-layout "Mean results" table.
SECTION_COLUMN = "Section"          # per-slice row label (the image file)
CORE_METRICS = ("Area", "Perimeter", "LGI", "Compactness")
SULCUS_CLASSES = ("Primary", "Secondary", "Tertiary", "Unclassified")
COUNT_METRICS = tuple(f"{c}SulciCount" for c in SULCUS_CLASSES)
MEAN_DEPTH_METRICS = tuple(f"{c}MeanDepth" for c in SULCUS_CLASSES)
# Individual per-sulcus depth columns, e.g. "Primary_depth_1", "Unclassified_depth_7".
DEPTH_VALUE_PATTERN = re.compile(r"^(Primary|Secondary|Tertiary|Unclassified)_depth_(\d+)$")

# ``process_on_images_batch`` writes "Batch_Allmarks.xlsx"; the master-report
# driver renames it "week{N}_{axis}_Batch_Allmarks.xlsx". Match both.
WORKBOOK_PATTERN = re.compile(
    r"(?:week(?P<week>\d+)_(?P<axis>axial|coronal|sagittal)_)?Batch_Allmarks\.xlsx$",
    re.IGNORECASE,
)
AXIS_NAMES = {"axial", "coronal", "sagittal"}
ANALYSIS_SHEET_NAME = "Analysis"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Analyze the report workbook of any process-image-batch run and write "
            "summary stats plus boxplots back into the same Excel file. Works for "
            "both the master-driver reports (measurements/…/week{N}_{axis}_Batch_"
            "Allmarks.xlsx) and a plain Batch_Allmarks.xlsx from the GUI batch."
        )
    )
    parser.add_argument(
        "inputs",
        nargs="*",
        type=Path,
        help=(
            "Workbook files and/or folders to analyze. A file is analyzed "
            "directly (any name); a folder is searched recursively with "
            "--report-glob. Defaults to --input-root when none are given."
        ),
    )
    parser.add_argument(
        "--input-root",
        type=Path,
        default=Path("measurements"),
        help="Fallback root searched when no positional inputs are given.",
    )
    parser.add_argument(
        "--axes",
        nargs="+",
        default=["axial", "coronal", "sagittal"],
        help="Axis names to include (only applied when the axis can be inferred).",
    )
    parser.add_argument(
        "--weeks",
        nargs="+",
        type=int,
        help="Optional weeks to include, for example: --weeks 24 25 26",
    )
    parser.add_argument(
        "--report-glob",
        default="*Batch_Allmarks.xlsx",
        help="Glob used below input-root to discover workbooks.",
    )
    parser.add_argument(
        "--sheet-name",
        default=ANALYSIS_SHEET_NAME,
        help="Worksheet name used for the generated analysis output.",
    )
    return parser.parse_args()


def infer_week_axis(path: Path) -> tuple[int | None, str | None]:
    """Best-effort ``(week, axis)`` for a workbook.

    First tries the ``week{N}_{axis}_`` filename prefix; failing that, scans the
    path components for an axis folder and a numeric week folder (the layout
    ``measurements/{week}/{section}/{axis}/…`` the master driver writes). Either
    element may be ``None`` when it cannot be determined (a plain
    ``Batch_Allmarks.xlsx`` dropped anywhere).
    """
    match = WORKBOOK_PATTERN.match(path.name)
    if match and match.group("week") is not None:
        return int(match.group("week")), match.group("axis").lower()

    week: int | None = None
    axis: str | None = None
    for part in path.parts:
        low = part.lower()
        if low in AXIS_NAMES:
            axis = low
        elif week is None and part.isdigit():
            week = int(part)
    return week, axis


def discover_workbooks(
    inputs: Iterable[Path], report_glob: str, axes: set[str], weeks: set[int] | None
) -> list[Path]:
    """Resolve *inputs* (files and/or folders) to a de-duplicated workbook list.

    A file input is trusted and analyzed directly regardless of its name (so a
    plain ``Batch_Allmarks.xlsx`` from the GUI batch works); a folder input is
    searched recursively with *report_glob* and the discovered names are filtered
    by the ``week{N}_{axis}_`` pattern and the ``--weeks`` / ``--axes`` options
    (applied only when week/axis can be inferred, so nothing is silently dropped).
    """
    paths: list[Path] = []
    seen: set[Path] = set()

    def _add(path: Path) -> None:
        resolved = path.resolve()
        if resolved not in seen:
            seen.add(resolved)
            paths.append(path)

    for item in inputs:
        if item.is_file():
            _add(item)  # explicit file — trust it
        elif item.is_dir():
            for path in sorted(item.rglob(report_glob)):
                if not WORKBOOK_PATTERN.search(path.name):
                    continue
                week, axis = infer_week_axis(path)
                if weeks is not None and week is not None and week not in weeks:
                    continue
                if axis is not None and axis not in axes:
                    continue
                _add(path)
        else:
            print(f"Warning: input not found, skipping: {item}")
    return paths


def parse_workbook(path: Path) -> tuple[pd.DataFrame, dict, dict]:
    """Read one batch workbook into ``(df, parameters, totals)``.

    ``df`` is the per-slice "Mean results" table with ``Section`` renamed to
    ``File``; ``parameters`` and ``totals`` are the spec-layout blocks.
    """
    parsed = read_results_sheet(str(path))
    rows = parsed.get("rows") or []
    if not rows:
        raise ValueError(
            f"{path.name}: no per-slice rows found — is this a "
            f"process-image-batch (spec-layout) workbook?"
        )
    df = pd.DataFrame(rows)
    if SECTION_COLUMN in df.columns:
        df = df.rename(columns={SECTION_COLUMN: "File"})
    # Keep only rows that name a slice (defensive; the reader already strips
    # the metadata/section blocks).
    df = df[df["File"].map(lambda v: isinstance(v, str) and v.strip() != "")].copy()

    week, axis = infer_week_axis(path)
    df["week"] = week
    df["axis"] = axis
    df["workbook"] = path.name
    return df.reset_index(drop=True), parsed.get("parameters") or {}, parsed.get("totals") or {}


def infer_depth_unit(parameters: dict) -> str:
    unit = parameters.get("Length unit")
    if isinstance(unit, str) and unit.strip():
        return unit.strip()
    spacing = parameters.get("Pixel spacing")
    if isinstance(spacing, str):
        match = re.search(r"\b([A-Za-z]+)\s*/\s*pixel", spacing)
        if match:
            return match.group(1)
    return "mm"


def existing_columns(df: pd.DataFrame, names: Iterable[str]) -> list[str]:
    return [name for name in names if name in df.columns]


def depth_value_columns(df: pd.DataFrame, class_name: str) -> list[str]:
    """Per-sulcus depth columns for one class, ordered ``_1``, ``_2``, …."""
    cols = []
    for col in df.columns:
        if not isinstance(col, str):
            continue
        match = DEPTH_VALUE_PATTERN.match(col)
        if match and match.group(1).lower() == class_name.lower():
            cols.append((int(match.group(2)), col))
    return [col for _, col in sorted(cols)]


def build_sulcus_value_table(df: pd.DataFrame) -> pd.DataFrame:
    """Long table of every individual sulcus depth: (slice_file, class, value, value_norm).

    Read straight from the ``{Class}_depth_{i}`` columns — no reconstruction. The
    adjacent ``{Class}_depth_{i}_norm`` column (each depth normalized to the
    deepest sulcus of its own image by
    :func:`functions.measurement_batch.process_on_images_batch`) is carried as
    ``value_norm``; ``NaN`` for older workbooks that lack the ``_norm`` columns.
    Always returns the fixed schema (empty when a workbook has no per-sulcus
    columns) so downstream lookups on ``sulcus_class`` never raise.
    """
    rows: list[dict[str, object]] = []
    for _, row in df.iterrows():
        for class_name in SULCUS_CLASSES:
            for col in depth_value_columns(df, class_name):
                value = row.get(col)
                if pd.notna(value):
                    norm = row.get(f"{col}_norm")
                    rows.append(
                        {
                            "slice_file": row["File"],
                            "sulcus_class": class_name.lower(),
                            "value": float(value),
                            "value_norm": float(norm) if pd.notna(norm) else math.nan,
                        }
                    )
    return pd.DataFrame(
        rows, columns=["slice_file", "sulcus_class", "value", "value_norm"]
    )


def metric_summary(df: pd.DataFrame, metrics: Iterable[str]) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for metric in metrics:
        if metric not in df.columns:
            continue
        values = pd.to_numeric(df[metric], errors="coerce").dropna()
        rows.append(
            {
                "metric": metric,
                "n": int(values.shape[0]),
                "mean": float(values.mean()) if not values.empty else math.nan,
                "std": float(values.std(ddof=1)) if len(values) > 1 else math.nan,
                "min": float(values.min()) if not values.empty else math.nan,
                "max": float(values.max()) if not values.empty else math.nan,
            }
        )
    return pd.DataFrame(rows)


def counts_per_slice_table(df: pd.DataFrame) -> pd.DataFrame:
    cols = ["File"] + existing_columns(df, COUNT_METRICS)
    out = df[cols].copy()
    for col in existing_columns(df, COUNT_METRICS):
        out[col] = pd.to_numeric(out[col], errors="coerce").round().astype("Int64")
    return out


def count_summary(counts_df: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for col in COUNT_METRICS:
        if col not in counts_df.columns:
            continue
        values = pd.to_numeric(counts_df[col], errors="coerce").dropna().astype(int)
        rows.append(
            {
                "metric": col,
                "n": int(values.shape[0]),
                "mean": float(values.mean()) if not values.empty else math.nan,
                "std": float(values.std(ddof=1)) if len(values) > 1 else math.nan,
                "min": int(values.min()) if not values.empty else pd.NA,
                "max": int(values.max()) if not values.empty else pd.NA,
            }
        )
    count_cols = existing_columns(counts_df, COUNT_METRICS)
    if count_cols:
        totals = (
            counts_df[count_cols]
            .apply(pd.to_numeric, errors="coerce")
            .sum(axis=1, min_count=1)
            .dropna()
        )
        rows.append(
            {
                "metric": "all_sulcus_count",
                "n": int(totals.shape[0]),
                "mean": float(totals.mean()) if not totals.empty else math.nan,
                "std": float(totals.std(ddof=1)) if len(totals) > 1 else math.nan,
                "min": int(totals.min()) if not totals.empty else pd.NA,
                "max": int(totals.max()) if not totals.empty else pd.NA,
            }
        )
    return pd.DataFrame(rows)


_DEPTH_SUMMARY_COLUMNS = ["metric", "n", "mean", "std", "min", "max"]


def _depth_stats_row(metric: str, values: pd.Series) -> dict[str, object]:
    """One summary row (mean/std/min/max) for *values*; ``std`` is ``NaN`` for n≤1."""
    v = pd.to_numeric(values, errors="coerce").dropna()
    if v.empty:
        return {"metric": metric, "n": 0, "mean": math.nan,
                "std": math.nan, "min": math.nan, "max": math.nan}
    return {
        "metric": metric,
        "n": int(v.shape[0]),
        "mean": float(v.mean()),
        "std": float(v.std(ddof=1)) if len(v) > 1 else math.nan,
        "min": float(v.min()),
        "max": float(v.max()),
    }


def per_class_depth_summary(sulcus_values_df: pd.DataFrame) -> pd.DataFrame:
    """Two rows per class: raw depth stats and, below it, the per-sulcus
    normalized-depth stats (each depth / its image's deepest sulcus)."""
    rows: list[dict[str, object]] = []
    for class_name in [name.lower() for name in SULCUS_CLASSES]:
        mask = sulcus_values_df["sulcus_class"] == class_name
        values = pd.to_numeric(
            sulcus_values_df.loc[mask, "value"], errors="coerce"
        ).dropna()
        if values.empty:
            continue
        rows.append(_depth_stats_row(f"{class_name}_sulcus_depth", values))
        rows.append(_depth_stats_row(
            f"{class_name}_sulcus_depth_normalized",
            sulcus_values_df.loc[mask, "value_norm"],
        ))
    return pd.DataFrame(rows, columns=_DEPTH_SUMMARY_COLUMNS)


def overall_depth_summary(sulcus_values_df: pd.DataFrame) -> pd.DataFrame:
    """Raw overall depth stats, followed by a separate normalized-depth row."""
    values = pd.to_numeric(sulcus_values_df["value"], errors="coerce").dropna()
    if values.empty:
        return pd.DataFrame(columns=_DEPTH_SUMMARY_COLUMNS)
    rows = [
        _depth_stats_row("all_sulcus_depths", values),
        _depth_stats_row("all_sulcus_depths_normalized", sulcus_values_df["value_norm"]),
    ]
    return pd.DataFrame(rows, columns=_DEPTH_SUMMARY_COLUMNS)


# ── Plotting ─────────────────────────────────────────────────────────────────
BOX_COLOR = "#9ecae1"
SCATTER_COLOR = "#2E6DA4"


def _save_figure(fig) -> io.BytesIO:
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=180)
    plt.close(fig)
    buf.seek(0)
    return buf


def _annotate_median(ax, position: float, median: float, fmt: str = "{:.3g}") -> None:
    ax.text(
        position,
        median,
        fmt.format(median),
        ha="center",
        va="bottom",
        fontsize=8,
        fontweight="bold",
        color="#1A3A5C",
    )


def plot_boxplot(values: pd.Series, title: str, ylabel: str) -> io.BytesIO | None:
    if plt is None:
        return None
    clean = pd.to_numeric(values, errors="coerce").dropna()
    if clean.empty:
        return None

    fig, ax = plt.subplots(figsize=(7, 4))
    bp = ax.boxplot(clean, patch_artist=True, boxprops={"facecolor": BOX_COLOR})
    ax.set_xticklabels([])
    ax.tick_params(axis="x", length=0)
    _annotate_median(ax, 1, float(bp["medians"][0].get_ydata()[0]))
    ax.set_title(title)
    ax.set_ylabel(ylabel)
    ax.grid(axis="y", linestyle=":", alpha=0.5)
    fig.tight_layout()
    return _save_figure(fig)


def plot_grouped_sulcus_boxplot(
    sulcus_values_df: pd.DataFrame, ylabel: str
) -> io.BytesIO | None:
    if plt is None:
        return None
    grouped = []
    labels = []
    for class_name in [name.lower() for name in SULCUS_CLASSES]:
        values = pd.to_numeric(
            sulcus_values_df.loc[
                sulcus_values_df["sulcus_class"] == class_name, "value"
            ],
            errors="coerce",
        ).dropna()
        if values.empty:
            continue
        grouped.append(values)
        labels.append(class_name)

    if not grouped:
        return None

    fig, ax = plt.subplots(figsize=(8.5, 4.5))
    bp = ax.boxplot(
        grouped, labels=labels, patch_artist=True, boxprops={"facecolor": BOX_COLOR}
    )
    for i, median_line in enumerate(bp["medians"]):
        _annotate_median(ax, i + 1, float(median_line.get_ydata()[0]))
    ax.set_title("Sulcus depth distributions by class")
    ax.set_ylabel(ylabel)
    ax.grid(axis="y", linestyle=":", alpha=0.5)
    fig.tight_layout()
    return _save_figure(fig)


def plot_grouped_count_boxplot(counts_df: pd.DataFrame) -> io.BytesIO | None:
    if plt is None:
        return None
    grouped = []
    labels = []
    for column_name, label in zip(
        COUNT_METRICS, [c.lower() for c in SULCUS_CLASSES]
    ):
        if column_name not in counts_df.columns:
            continue
        values = pd.to_numeric(counts_df[column_name], errors="coerce").dropna()
        if values.empty:
            continue
        grouped.append(values)
        labels.append(label)

    if not grouped:
        return None

    fig, ax = plt.subplots(figsize=(9.5, 5.0))
    positions = list(range(1, len(grouped) + 1))
    bp = ax.boxplot(
        grouped, labels=labels, positions=positions, patch_artist=True,
        boxprops={"facecolor": BOX_COLOR},
    )
    for i, (pos, values) in enumerate(zip(positions, grouped)):
        offsets = np.linspace(-0.08, 0.08, len(values)) if len(values) > 1 else np.array([0.0])
        ax.scatter(
            pos + offsets, values, s=20, color=SCATTER_COLOR, alpha=0.7,
            edgecolors="none", zorder=3,
        )
        _annotate_median(ax, pos, float(bp["medians"][i].get_ydata()[0]), fmt="{:.0f}")
    ax.set_title("Sulcus count distributions by fold class")
    ax.set_ylabel("count per slice")
    max_count = max(int(values.max()) for values in grouped)
    ax.set_ylim(bottom=-0.1, top=max_count + 0.35)
    ax.set_yticks(list(range(0, max_count + 1)))
    ax.grid(axis="y", linestyle=":", alpha=0.5)
    fig.tight_layout()
    return _save_figure(fig)


# ── Excel styling / writing ──────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="2E6DA4")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
SECTION_FILL = PatternFill("solid", fgColor="D9E8F5")
SECTION_FONT = Font(bold=True, color="1A3A5C", size=11)
ROW_FILL_ODD = PatternFill("solid", fgColor="FFFFFF")
ROW_FILL_EVEN = PatternFill("solid", fgColor="EEF4FB")
BORDER_SIDE = Side(style="thin", color="B0C4D8")
CELL_BORDER = Border(
    left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE
)


def _style_cell(cell, fill=None, font=None, border=True, align_center=False) -> None:
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if border:
        cell.border = CELL_BORDER
    if align_center:
        cell.alignment = Alignment(horizontal="center")


def _autofit_columns(ws) -> None:
    col_widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            text_len = len(str(cell.value))
            col_widths[cell.column] = max(col_widths.get(cell.column, 0), text_len)
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width + 3, 50)


def _set_wrapped_row_height(ws, row: int, columns: Iterable[int], base_height: int = 18) -> None:
    max_lines = 1
    for col in columns:
        cell = ws.cell(row=row, column=col)
        if cell.value is None:
            continue
        text = str(cell.value)
        width = ws.column_dimensions[get_column_letter(col)].width or 10
        approx_chars = max(12, int(width * 1.15))
        wrapped_lines = sum(
            max(1, math.ceil(len(part) / approx_chars)) for part in text.splitlines() or [""]
        )
        max_lines = max(max_lines, wrapped_lines)
    ws.row_dimensions[row].height = base_height * max_lines


def normalize_excel_value(value: object) -> object:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    return value


def write_dataframe(ws, start_row: int, start_col: int, title: str, df: pd.DataFrame) -> tuple[int, int]:
    """Write a titled table; returns (next_free_row, rightmost_column_used)."""
    num_cols = max(len(df.columns), 1)
    end_col = start_col + num_cols - 1

    title_cell = ws.cell(row=start_row, column=start_col, value=title)
    _style_cell(title_cell, fill=SECTION_FILL, font=SECTION_FONT, border=False)
    if num_cols > 1:
        ws.merge_cells(
            start_row=start_row, start_column=start_col,
            end_row=start_row, end_column=end_col,
        )

    if df.empty:
        ws.cell(row=start_row + 1, column=start_col, value="No data")
        return start_row + 3, end_col

    header_row = start_row + 1
    for offset, col_name in enumerate(df.columns):
        cell = ws.cell(row=header_row, column=start_col + offset, value=str(col_name))
        _style_cell(cell, fill=HEADER_FILL, font=HEADER_FONT)

    for row_offset, row in enumerate(df.itertuples(index=False), start=1):
        fill = ROW_FILL_EVEN if row_offset % 2 == 0 else ROW_FILL_ODD
        for col_offset, value in enumerate(row):
            cell = ws.cell(
                row=header_row + row_offset,
                column=start_col + col_offset,
                value=normalize_excel_value(value),
            )
            _style_cell(cell, fill=fill)

    return header_row + len(df) + 2, end_col


def replace_analysis_sheet(workbook_path: Path, sheet_name: str):
    wb = load_workbook(workbook_path)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)
    return wb, ws


def add_image(ws, image_bytes: io.BytesIO | None, anchor: str, width: int = 520, height: int = 360) -> None:
    if image_bytes is None or XLImage is None:
        return
    img = XLImage(image_bytes)
    img.width = width
    img.height = height
    ws.add_image(img, anchor)


def analyze_workbook(path: Path | str, sheet_name: str = ANALYSIS_SHEET_NAME) -> None:
    path = Path(path)
    df, parameters, totals = parse_workbook(path)
    depth_unit = infer_depth_unit(parameters)

    for col in existing_columns(df, CORE_METRICS):
        df[col] = pd.to_numeric(df[col], errors="coerce")

    counts_df = counts_per_slice_table(df)
    core_summary_df = metric_summary(df, CORE_METRICS)
    count_summary_df = count_summary(counts_df)
    sulcus_values_df = build_sulcus_value_table(df)
    has_sulcus_values = not sulcus_values_df.empty
    per_class_depth_df = per_class_depth_summary(sulcus_values_df)
    overall_depth_df = overall_depth_summary(sulcus_values_df)

    wb, ws = replace_analysis_sheet(path, sheet_name)
    week, axis = infer_week_axis(path)

    # ── Metadata info card ───────────────────────────────────────────────────
    CARD_TITLE_FILL = PatternFill("solid", fgColor="1A3A5C")
    CARD_TITLE_FONT = Font(bold=True, color="FFFFFF", size=12)
    CARD_LABEL_FILL = PatternFill("solid", fgColor="D9E8F5")
    CARD_LABEL_FONT = Font(bold=True, color="1A3A5C", size=10)
    CARD_VALUE_FILL = PatternFill("solid", fgColor="EEF4FB")
    CARD_VALUE_FONT = Font(color="1A1A1A", size=10)

    title_cell = ws["A1"]
    title_cell.value = "Workbook Analysis"
    title_cell.fill = CARD_TITLE_FILL
    title_cell.font = CARD_TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 22

    card_rows = [
        ("Workbook", path.name),
        ("Week", week if week is not None else "—"),
        ("Axis", axis if axis is not None else "—"),
        ("Slice rows analyzed", len(df)),
        ("Total sulci count", totals.get("Total sulci count", "—")),
    ]
    classes_present = sorted(
        {c for c in sulcus_values_df["sulcus_class"].unique()}
    ) if has_sulcus_values else []
    note_parts = []
    if has_sulcus_values:
        note_parts.append(
            "Per-class sulcus depth plots use the real per-sulcus values "
            f"({', '.join(classes_present)}) read straight from the "
            "{Class}_depth_i columns."
        )
        if classes_present == ["unclassified"]:
            note_parts.append(
                "Only the 'unclassified' class is present, so these are cropped "
                "sub-slices measured with the fixed-depth sulci rule (no "
                "primary/secondary/tertiary breakdown)."
            )
    else:
        note_parts.append(
            "This workbook has no per-sulcus depth columns; only counts and core "
            "metrics are summarized."
        )
    if MATPLOTLIB_IMPORT_ERROR is not None:
        note_parts.append(
            "Embedded plot images were skipped because matplotlib was unavailable."
        )
    note_text = " ".join(note_parts)

    for idx, (label, value) in enumerate(card_rows, start=2):
        lbl = ws.cell(row=idx, column=1, value=label)
        lbl.fill = CARD_LABEL_FILL
        lbl.font = CARD_LABEL_FONT
        lbl.border = CELL_BORDER
        val = ws.cell(row=idx, column=2, value=normalize_excel_value(value))
        val.fill = CARD_VALUE_FILL
        val.font = CARD_VALUE_FONT
        val.border = CELL_BORDER

    note_row = 2 + len(card_rows)
    note_label = ws.cell(row=note_row, column=1, value="Note")
    note_label.fill = CARD_LABEL_FILL
    note_label.font = CARD_LABEL_FONT
    note_label.border = CELL_BORDER
    note_value = ws.cell(row=note_row, column=2, value=note_text)
    note_value.fill = CARD_VALUE_FILL
    note_value.font = CARD_VALUE_FONT
    note_value.border = CELL_BORDER
    note_value.alignment = Alignment(wrap_text=True)

    # ── Data tables ──────────────────────────────────────────────────────────
    next_row = note_row + 2
    max_data_col = 1

    def _emit(title: str, table: pd.DataFrame) -> None:
        nonlocal next_row, max_data_col
        next_row, used_col = write_dataframe(ws, next_row, 1, title, table)
        max_data_col = max(max_data_col, used_col)

    _emit("Core Metric Summary", core_summary_df)
    _emit("Sulcus Counts Per Slice", counts_df)
    _emit("Sulcus Count Summary", count_summary_df)
    if has_sulcus_values:
        _emit("Per-Class Sulcus Depth Summary", per_class_depth_df)
        _emit("Overall Sulcus Depth Summary", overall_depth_df)

    # ── Plot anchors: placed 2 columns right of the widest table ─────────────
    plot_col_1 = max_data_col + 2          # core metric plots
    plot_col_2 = plot_col_1 + 9            # per-class sulcus depth plots
    plot_col_3 = plot_col_2 + 9            # grouped summary plots

    def _col_anchor(col: int, row: int) -> str:
        return f"{get_column_letter(col)}{row}"

    PLOT_ROW_STEP = 22
    for i, metric in enumerate(CORE_METRICS):
        if metric not in df.columns:
            continue
        add_image(
            ws,
            plot_boxplot(df[metric], f"{metric} distribution", metric),
            _col_anchor(plot_col_1, 2 + i * PLOT_ROW_STEP),
        )

    if has_sulcus_values:
        for i, class_name in enumerate([c.lower() for c in SULCUS_CLASSES]):
            class_values = sulcus_values_df.loc[
                sulcus_values_df["sulcus_class"] == class_name, "value"
            ]
            if class_values.empty:
                continue
            add_image(
                ws,
                plot_boxplot(
                    class_values,
                    f"{class_name.capitalize()} sulcus depth distribution",
                    f"depth ({depth_unit})",
                ),
                _col_anchor(plot_col_2, 2 + i * PLOT_ROW_STEP),
            )
        add_image(
            ws,
            plot_grouped_sulcus_boxplot(sulcus_values_df, f"depth ({depth_unit})"),
            _col_anchor(plot_col_3, 2),
            width=620,
            height=360,
        )
    add_image(
        ws,
        plot_grouped_count_boxplot(counts_df),
        _col_anchor(plot_col_3, 2 + PLOT_ROW_STEP),
        width=680,
        height=400,
    )

    _autofit_columns(ws)
    _set_wrapped_row_height(ws, note_row, [1, 2])
    wb.save(path)


def main() -> int:
    args = parse_args()
    axes = {axis.strip().lower() for axis in args.axes}
    weeks = set(args.weeks) if args.weeks else None

    inputs = list(args.inputs) if args.inputs else [args.input_root]
    report_paths = discover_workbooks(inputs, args.report_glob, axes, weeks)
    if not report_paths:
        raise SystemExit(
            "No matching workbooks found in: "
            + ", ".join(str(item) for item in inputs)
        )

    for path in report_paths:
        analyze_workbook(path, args.sheet_name)
        print(f"Updated workbook: {path}")

    if MATPLOTLIB_IMPORT_ERROR is not None:
        print(
            "Warning: matplotlib could not be imported, so analysis sheets were "
            "written without embedded plot images."
        )
        print(f"Import error: {MATPLOTLIB_IMPORT_ERROR}")

    print(f"Analyzed {len(report_paths)} workbooks.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
