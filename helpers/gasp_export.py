"""Build the per-run GASP results folder.

Writes an Excel summary (header block + one row per gestational week),
copies the source measured image, and is intended to be called once per
similarity-profile run.  The folder it populates is the temporary results
directory whose lifetime is managed by :class:`MainWindow`; ``Save Data``
later copies the whole folder to a permanent location.
"""

from __future__ import annotations

import os
import shutil
from datetime import datetime
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from helpers.gestational_week_profile import (
    GASPSummary,
    GestationalWeekProfile,
)


MEASURED_METRIC_KEYS: tuple[str, ...] = (
    "Area", "Perimeter", "LGI", "Compactness",
    "PrimarySulciCount", "SecondarySulciCount",
    "TertiarySulciCount", "UnclassifiedSulciCount",
    "PrimaryMeanDepth", "SecondaryMeanDepth",
    "TertiaryMeanDepth", "UnclassifiedMeanDepth",
)

REFERENCE_METRIC_FIELDS: tuple[str, ...] = (
    "area", "perimeter", "lgi", "compactness",
    "primary_count", "secondary_count", "tertiary_count",
    "primary_sulcus_values", "secondary_sulcus_values",
    "tertiary_sulcus_values",
)


def export_gasp_results(
    out_dir: str,
    *,
    project_name: str,
    source_path: str | None,
    measured: dict,
    summary: GASPSummary,
    summary_alt: GASPSummary | None,
    registry: GestationalWeekProfile,
    axis: str,
    params: dict[str, Any],
    weights: dict[str, float] | None = None,
    excel_filename: str = "gasp_results.xlsx",
) -> dict[str, str]:
    """Populate *out_dir* with the Excel summary and source-image copy.

    Returns a dict mapping artifact name to path.
    """
    os.makedirs(out_dir, exist_ok=True)
    artifacts: dict[str, str] = {}

    src_copy = _copy_source_image(out_dir, source_path)
    if src_copy:
        artifacts["source_image"] = src_copy

    excel_path = os.path.join(out_dir, excel_filename)
    _write_excel(
        excel_path,
        project_name=project_name,
        source_path=source_path,
        measured=measured,
        summary=summary,
        summary_alt=summary_alt,
        registry=registry,
        axis=axis,
        params=params,
        weights=weights,
    )
    artifacts["excel"] = excel_path
    return artifacts


def _copy_source_image(out_dir: str, source_path: str | None) -> str | None:
    if not source_path or not os.path.isfile(source_path):
        return None
    dst = os.path.join(out_dir, "source_" + os.path.basename(source_path))
    if os.path.abspath(source_path) == os.path.abspath(dst):
        return dst
    shutil.copy2(source_path, dst)
    return dst


def _write_excel(
    excel_path: str,
    *,
    project_name: str,
    source_path: str | None,
    measured: dict,
    summary: GASPSummary,
    summary_alt: GASPSummary | None,
    registry: GestationalWeekProfile,
    axis: str,
    params: dict[str, Any],
    weights: dict[str, float] | None,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "GASP Results"

    title_font = Font(bold=True, size=14, color="FFFFFF")
    title_fill = PatternFill("solid", fgColor="1F6FA5")
    section_font = Font(bold=True, size=11, color="FFFFFF")
    section_fill = PatternFill("solid", fgColor="2C3E50")
    bold = Font(bold=True)
    table_header_font = Font(bold=True, color="FFFFFF")
    table_header_fill = PatternFill("solid", fgColor="3498DB")
    centered = Alignment(horizontal="center")

    row = 1
    title_cell = ws.cell(row=row, column=1,
                         value="Gestational Age Similarity Profile (GASP)")
    title_cell.font = title_font
    title_cell.fill = title_fill
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 2

    header_items = _build_header_items(
        project_name=project_name,
        source_path=source_path,
        axis=axis,
        params=params,
        measured=measured,
        summary=summary,
        summary_alt=summary_alt,
        weights=weights,
    )

    for item in header_items:
        kind = item[0]
        if kind == "section":
            cell = ws.cell(row=row, column=1, value=item[1])
            cell.font = section_font
            cell.fill = section_fill
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=6)
            row += 1
        elif kind == "kv":
            ws.cell(row=row, column=1, value=item[1]).font = bold
            ws.cell(row=row, column=2, value=_to_cell_value(item[2]))
            row += 1
        elif kind == "blank":
            row += 1

    row += 1

    per_week_df = pd.DataFrame(
        _build_per_week_rows(summary, summary_alt, registry, axis)
    )
    columns = list(per_week_df.columns)
    for ci, col in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=ci, value=col)
        cell.font = table_header_font
        cell.fill = table_header_fill
        cell.alignment = centered
    row += 1
    for _, record in per_week_df.iterrows():
        for ci, col in enumerate(columns, start=1):
            ws.cell(row=row, column=ci, value=_to_cell_value(record[col]))
        row += 1

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 42
    for ci in range(3, max(len(columns) + 1, 7)):
        ws.column_dimensions[get_column_letter(ci)].width = 16

    wb.save(excel_path)


def _build_header_items(
    *,
    project_name: str,
    source_path: str | None,
    axis: str,
    params: dict[str, Any],
    measured: dict,
    summary: GASPSummary,
    summary_alt: GASPSummary | None,
    weights: dict[str, float] | None,
) -> list[tuple]:
    items: list[tuple] = []

    items.append(("section", "Run Information"))
    items.append(("kv", "Project / Result Name", project_name))
    items.append(("kv", "Source File", source_path or ""))
    items.append(("kv", "Timestamp",
                  datetime.now().isoformat(timespec="seconds")))
    items.append(("kv", "Axis", axis))
    items.append(("blank",))

    items.append(("section", "Analysis Parameters"))
    items.append(("kv", "GASP Method", params.get("method_banner", "")))
    items.append(("kv", "Range Penalty (lambda)", params.get("range_penalty")))
    items.append(("kv", "OOR Beta (beta)", params.get("oor_beta")))
    items.append(("kv", "Apply Range Penalty", params.get("apply_penalty")))
    items.append(("kv", "Weighted Global", params.get("weighted_global")))
    items.append(("kv", "Kernel Size (mm)", params.get("kernel_size_mm", params.get("kernel_size"))))
    items.append(("kv", "Kernel Size (px)", params.get("kernel_size_px")))
    pxsize = params.get("pixel_size")
    pxunit = params.get("pixel_size_units") or ""
    pxstr = (f"{pxsize} {pxunit}/pixel".strip()
             if pxsize is not None else "")
    items.append(("kv", "Pixel Spacing", pxstr))
    items.append(("kv", "Length Unit", params.get("length_unit")))
    items.append(("kv", "Filtered Threshold (mm²)", params.get("filtered_threshold")))
    items.append(("kv", "Contour Mode", params.get("contour_mode")))
    items.append(("blank",))

    items.append(("section", "Measured Hallmark Values"))
    for k in MEASURED_METRIC_KEYS:
        v = measured.get(k)
        if v is None:
            continue
        items.append(("kv", k, v))
    items.append(("blank",))

    items.append(("section", "GASP Summary"))
    items.append(("kv", "Best Week", summary.best_week))
    items.append(("kv", "Max GASP", summary.max_gasp))
    items.append(("kv", "Estimated GA (weeks)", summary.estimated_ga))
    items.append(("kv", "Confidence", summary.confidence))
    if summary_alt is not None:
        items.append(("kv", "Best Week (Global Distance)",
                      summary_alt.best_week))
        items.append(("kv", "Max GASP (Global Distance)",
                      summary_alt.max_gasp))
        items.append(("kv", "Estimated GA (Global Distance)",
                      summary_alt.estimated_ga))
        items.append(("kv", "Confidence (Global Distance)",
                      summary_alt.confidence))
    items.append(("blank",))

    if weights:
        items.append(("section", "Metric Weights"))
        for k, v in weights.items():
            items.append(("kv", k, v))
        items.append(("blank",))

    return items


def _build_per_week_rows(
    summary: GASPSummary,
    summary_alt: GASPSummary | None,
    registry: GestationalWeekProfile,
    axis: str,
) -> list[dict[str, Any]]:
    primary_by_week = {r.week: r for r in summary.results}
    alt_by_week = {r.week: r for r in (summary_alt.results
                                       if summary_alt is not None else [])}
    weeks = sorted(set(primary_by_week) | set(alt_by_week))
    has_alt = bool(alt_by_week)

    rows: list[dict[str, Any]] = []
    for week in weeks:
        ref = registry.get(week, axis)
        primary = primary_by_week.get(week)
        alt = alt_by_week.get(week)
        chosen = primary or alt
        row: dict[str, Any] = {"Week": week}
        if has_alt:
            row["GASP (Gaussian)"] = primary.gasp if primary else None
            row["GASP (Global Distance)"] = alt.gasp if alt else None
        else:
            row["GASP"] = chosen.gasp if chosen else None

        for f in REFERENCE_METRIC_FIELDS:
            row[f"sim_{f}"] = chosen.per_metric.get(f) if chosen else None
            row[f"z_{f}"] = chosen.z_scores.get(f) if chosen else None
            oor = chosen.out_of_range.get(f) if chosen else None
            row[f"oor_{f}"] = (bool(oor) if oor is not None else None)

        if ref is not None:
            for f in REFERENCE_METRIC_FIELDS:
                stats = getattr(ref, f, None)
                row[f"ref_{f}_mean"] = stats.mean if stats else None
                row[f"ref_{f}_std"] = stats.std if stats else None
        rows.append(row)
    return rows


def _to_cell_value(value: Any) -> Any:
    """Coerce a value into something openpyxl is happy to store."""
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float, str)):
        try:
            if isinstance(value, float) and pd.isna(value):
                return None
        except Exception:
            pass
        return value
    return str(value)
